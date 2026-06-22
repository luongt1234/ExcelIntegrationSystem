"""
controllers/people_controller.py
==================================
PeopleController – Điều phối toàn bộ flow Gộp Hồ Sơ Người.

Trách nhiệm:
    • Lắng nghe sự kiện từ View (qua command được gán vào nút).
    • Gọi Service (DataProcessor / LocalLearningManager) trong thread nền.
    • Cập nhật View qua .after(0, lambda: ...) – tuân thủ Thread-Safety.
    • KHÔNG vẽ widget, KHÔNG truy cập _var của View.

Nguyên tắc bất biến:
    1. Tất cả gọi Service chạy trong threading.Thread(daemon=True).
    2. Mọi callback cập nhật UI: self.window.after(0, lambda: ...).
    3. Lấy dữ liệu từ View chỉ qua DTO Getter (get_*).
"""

from __future__ import annotations

import os
import copy
import threading
from typing import TYPE_CHECKING, Dict, List, Optional, Tuple

import pandas as pd
from PyQt6.QtWidgets import (
    QFileDialog, QMessageBox,
    QDialog, QDialogButtonBox, QVBoxLayout, QLabel, QComboBox,
)
from PyQt6.QtCore import QObject, pyqtSignal

if TYPE_CHECKING:
    from views.main_window import MainWindow
    from services.data_processor import DataProcessor
    from services.learning_manager import LocalLearningManager


class PeopleController:
    """
    Controller cho flow Gộp Hồ Sơ Người.

    Được khởi tạo tại main.py với DI:
        ctrl = PeopleController(window, processor, lm)
    """

    STANDARD_COLUMNS = [
        "Họ và Tên", "Ngày Sinh", "Số CCCD/ID",
        "Số Điện Thoại", "Email", "Địa Chỉ",
        "Vị trí", "Đơn vị", "Ngày nộp",
        "Tình trạng", "Văn bằng", "Ghi Chú",
    ]

    def __init__(
        self,
        window: "MainWindow",
        processor: "DataProcessor",
        lm: "LocalLearningManager",
        app_controller=None,     # tham chiếu ngược để chuyển màn hình
    ) -> None:
        self.window = window
        self.processor = processor
        self.lm = lm
        self.app_ctrl = app_controller

        # ── State nội bộ ──────────────────────────────────────────────
        self._selected_files: List[str] = []
        self._merge_mode: int = 1
        self._file_structures: Dict[str, dict] = {}
        self._scan_results: Dict[str, dict] = {}
        self._saved_profiles: Dict[str, Optional[dict]] = {}

        self._raw_combined_df: Optional[pd.DataFrame] = None
        self._suspect_pairs: List[Tuple[int, int]] = []
        self._current_pair_index: int = 0
        self._rows_to_delete: set = set()
        self._merge_decisions: List[Tuple[int, int]] = []
        self._sort_config: dict = {"col": None, "ascending": True}
        self._left_join_result_df: Optional[pd.DataFrame] = None
        self._key_pair_config: List[Dict] = []
        self._clean_result_df: Optional[pd.DataFrame] = None

    # ==================================================================
    # 1. TỪ HOME: CHỌN FILE VÀ CẤU HÌNH
    # ==================================================================
    def handle_browse_files(self) -> None:
        """Mở dialog chọn file, cập nhật HomeScreen."""
        files, _ = QFileDialog.getOpenFileNames(
            self.window,
            "Chọn file Excel cần gộp",
            "",
            "Excel Files (*.xlsx *.xls)",
        )
        if not files:
            return
        self._selected_files = list(files)
        # Yêu cầu HomeScreen cập nhật UI — Controller không vẽ
        if hasattr(self, "_home_screen") and self._home_screen:
            self._home_screen.set_selected_files(self._selected_files)

    def handle_proceed_to_structure(self, home_screen) -> None:
        """
        Từ HomeScreen → chuyển sang StructureConfirmScreen.
        Không còn yêu cầu khai báo chế độ trước — chế độ được quyết định
        tại màn hình D&D sau khi xác nhận cấu trúc.
        """
        self._selected_files = home_screen.get_selected_files()

        if not self._selected_files:
            QMessageBox.warning(self.window, "Chưa chọn file", "Vui lòng chọn ít nhất 1 file Excel.")
            return

        if self._merge_mode == 2:
            # Nếu là chế độ Bổ sung, cần chọn File Gốc/Phụ trước khi quét
            if self.app_ctrl:
                self.app_ctrl.show_source_file_select()
        else:
            # Các chế độ khác quét cấu trúc luôn
            self._scan_all_files()

    def _scan_all_files(self) -> None:
        """Quét cấu trúc tất cả file, sau đó hiển thị StructureConfirmScreen."""
        self._scan_results = {}
        self._saved_profiles = {}
        self._file_structures = {}

        for path in self._selected_files:
            saved_profile = self.lm.find_similar_structure_profile(
                os.path.basename(path)
            )
            self._saved_profiles[path] = saved_profile
            
            sheet_name_to_scan = saved_profile.get("sheet_name") if saved_profile else None

            try:
                scan = self.processor.scan_excel_structure(path, sheet_name=sheet_name_to_scan, max_rows=15)
            except Exception as e:
                from PyQt6.QtWidgets import QMessageBox
                QMessageBox.warning(
                    self.window, 
                    "Lỗi đọc file", 
                    f"Không thể tải file {os.path.basename(path)}.\n\n{str(e)}"
                )
                self.window.update_status("Lỗi tải file")
                return
            self._scan_results[path] = copy.deepcopy(scan)

        # Hiển thị màn hình xác nhận cấu trúc (phải chạy ở main thread)
        self._goto_structure_confirm()

    def _goto_structure_confirm(self) -> None:
        """Chuyển sang StructureConfirmScreen."""
        if not self.app_ctrl:
            return

        # Tạo callback TRƯỚC khi tạo màn hình để tránh race condition
        rescan_cb = self._make_rescan_callback()

        # Truyền merge_mode và rescan_callback vào show_structure_confirm
        # AppController mới nhận rescan_callback; AppController cũ dùng TypeError fallback
        try:
            self.app_ctrl.show_structure_confirm(
                file_paths=self._selected_files,
                scan_results=self._scan_results,
                saved_profiles=self._saved_profiles,
                merge_mode=self._merge_mode,
                pre_confirmed_structures=dict(self._file_structures),
                rescan_callback=rescan_cb,
            )
        except TypeError:
            # AppController cũ không nhận rescan_callback → thử không có
            try:
                self.app_ctrl.show_structure_confirm(
                    file_paths=self._selected_files,
                    scan_results=self._scan_results,
                    saved_profiles=self._saved_profiles,
                    merge_mode=self._merge_mode,
                    pre_confirmed_structures=dict(self._file_structures),
                )
            except TypeError:
                self.app_ctrl.show_structure_confirm(
                    file_paths=self._selected_files,
                    scan_results=self._scan_results,
                    saved_profiles=self._saved_profiles,
                )

        # Dù AppController cũ/mới, luôn gán lại callback vào screen sau khi tạo
        # (đảm bảo screen luôn có callback dù constructor không nhận được)
        screen = None
        for attr in ("_current_screen", "current_screen", "_active_screen"):
            screen = getattr(self.app_ctrl, attr, None)
            if screen is not None:
                break
        if screen is None:
            get_screen = getattr(self.app_ctrl, "get_current_screen", None)
            if callable(get_screen):
                screen = get_screen()
        if screen is not None and hasattr(screen, "_rescan_callback"):
            screen._rescan_callback = rescan_cb

    def _make_rescan_callback(self):
        """
        Trả về hàm callback để StructureConfirmScreen gọi khi user đổi sheet / bấm Dò lại.
        Callback chạy ở main thread (scan nhẹ, không cần Thread).
        """
        def _rescan(file_path: str, sheet_name: str) -> dict:
            """Quét lại cấu trúc file với sheet chỉ định."""
            try:
                scan = self.processor.scan_excel_structure(
                    file_path, sheet_name=sheet_name, max_rows=15
                )
                self._scan_results[file_path] = copy.deepcopy(scan)
            except Exception as e:
                raise RuntimeError(f"scan_excel_structure thất bại: {e}") from e
            return scan
        return _rescan

    # ==================================================================
    # 2. BƯỚC 0: XÁC NHẬN CẤU TRÚC
    # ==================================================================
    def handle_structure_confirmed(self, structure_screen) -> None:
        """
        Từ StructureConfirmScreen → ModeSelectionScreen.
        """
        self._file_structures = structure_screen.get_file_structures()

        if not structure_screen.is_all_confirmed():
            QMessageBox.warning(
                self.window,
                "Chưa xác nhận đủ",
                "Vui lòng xác nhận cấu trúc tất cả các file trước khi tiếp tục.",
            )
            return

        # Lưu profile vào learning manager
        for path, struct in self._file_structures.items():
            profile_key = f"structure::{os.path.basename(path)}"
            self.lm.save_structure_profile(
                profile_key, os.path.basename(path), struct
            )

        if self.app_ctrl:
            if self._merge_mode == 3:
                self.app_ctrl.show_data_cleaner_screen(self._file_structures)
            elif self._merge_mode == 2:
                self._goto_drag_drop_mapping()
            else:
                self.app_ctrl.show_unified_key_config(self._file_structures, merge_mode=self._merge_mode)

    def _goto_drag_drop_mapping(self) -> None:
        """Thu thập danh sách cột từ Master và Aux để tạo giao diện Kéo-thả."""
        import re
        file_col_map = {}
        for path in self._selected_files:
            struct = self._file_structures.get(path, {})
            mapping = struct.get("column_mapping", {})
            selected = struct.get("selected_columns", [])
            final_cols = []
            for c in selected:
                c_norm = re.sub(r'\s+', ' ', str(c)).strip()
                mapped = mapping.get(c_norm) or mapping.get(c, "Bỏ qua (Ignore)")
                if mapped and mapped != "Bỏ qua (Ignore)":
                    final_cols.append(mapped)
                else:
                    final_cols.append(c_norm)
            unique_cols = []
            for c in final_cols:
                if c not in unique_cols:
                    unique_cols.append(c)
            file_col_map[path] = unique_cols

        self.app_ctrl.show_drag_drop_mapping(
            self._selected_files,
            file_col_map,
            pre_confirmed_structures=self._file_structures,
        )


    def handle_start_cleaning(self, screen) -> None:
        """Thực thi Dọn dẹp 1 File (Chế độ 3) → chuyển sang màn hình xem trước kết quả."""
        categorical_cols = screen.get_selected_columns()
        mode = screen.get_clean_mode()

        if not self._selected_files:
            return

        file_path = self._selected_files[0]
        struct = self._file_structures.get(file_path, {})

        self.window.update_status(message="⏳  Đang chuẩn hóa dữ liệu...", progress=50)

        try:
            cleaned_df = self.processor.normalize_categorical_columns(
                struct=struct,
                categorical_cols=categorical_cols,
                mode=mode,
                standard_columns=self.STANDARD_COLUMNS
            )

            # Lưu kết quả vào state để DataCleanerResultScreen hiển thị xem trước,
            # và để handle_export_clean dùng khi người dùng bấm "Xuất Excel".
            self._clean_result_df = cleaned_df
            self.window.update_status(message="✅  Đã chuẩn hóa xong, đang mở xem trước...", progress=100)

            if self.app_ctrl:
                self.app_ctrl.show_data_cleaner_result(cleaned_df)
            self.window.update_status("Sẵn sàng")
        except Exception as e:
            QMessageBox.critical(self.window, "Lỗi", f"Có lỗi xảy ra: {e}")
            self.window.update_status("Lỗi chuẩn hóa")

    def handle_export_clean(self) -> None:
        """
        Xuất kết quả Dọn dẹp & Chuẩn hóa (Chế độ 3) ra file Excel.
        Luôn xuất toàn bộ self._clean_result_df (không bị cắt theo giới hạn xem trước).
        """
        import os as _os
        import platform
        import subprocess

        if self._clean_result_df is None or self._clean_result_df.empty:
            QMessageBox.warning(self.window, "Không có dữ liệu", "Chưa có dữ liệu để xuất.")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self.window,
            "Lưu file Excel đã dọn dẹp",
            "Cleaned_Data.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not save_path:
            return

        if not save_path.endswith(".xlsx"):
            save_path += ".xlsx"

        try:
            self._export_formatted_excel(self._clean_result_df, save_path)
            self.window.update_status(message="✅ Đã lưu file", progress=100)

            reply = QMessageBox.question(
                self.window,
                "Hoàn tất",
                "Xuất file thành công! Bạn có muốn mở file ngay không?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                if platform.system() == "Darwin":
                    subprocess.call(("open", save_path))
                elif platform.system() == "Windows":
                    _os.startfile(save_path)
                else:
                    subprocess.call(("xdg-open", save_path))

            if self.app_ctrl:
                self.app_ctrl.show_home()
        except Exception as e:
            QMessageBox.critical(self.window, "Lỗi xuất file", f"Có lỗi xảy ra: {e}")
            self.window.update_status("Lỗi xuất file")

    def _run_left_join_with_mapping(
        self,
        key_columns: List[str],
        rename_map: Dict[str, str],
    ) -> None:
        """
        Chạy Sequential Left Join với mapping từ D&D.
        key_columns – tên cột (File Gốc) dùng làm khóa đối chiếu.
        rename_map  – {aux_col: master_col} để đổi tên cột phụ trước join.
        """
        file_structures_snapshot = dict(self._file_structures)
        self.window.update_status(message="⏳  Đang gộp chọn lọc...", progress=5)

        class JoinSignals(QObject):
            progress = pyqtSignal(str, float)
            finished = pyqtSignal(object)
            error    = pyqtSignal(str)

        signals = JoinSignals()
        signals.progress.connect(lambda m, p: self.window.update_status(message=m, progress=p))
        signals.finished.connect(
            lambda df: self.app_ctrl.show_left_join_result(df, key_columns)
        )
        signals.error.connect(
            lambda err: QMessageBox.critical(self.window, "Lỗi xử lý", f"Gộp chọn lọc thất bại:\n{err}")
        )

        def progress_cb(msg: str, pct: float) -> None:
            signals.progress.emit(msg, pct)

        def worker() -> None:
            try:
                if not rename_map:
                    result_df = self.processor.run_sequential_left_join(
                        file_paths=self._selected_files,
                        key_columns=key_columns,
                        file_structures=file_structures_snapshot,
                        progress_callback=progress_cb,
                    )
                else:
                    result_df = self._run_rename_join(
                        key_columns, rename_map, progress_cb,
                        file_structures=file_structures_snapshot,
                    )
                self._left_join_result_df = result_df
                signals.finished.emit(result_df)
            except Exception as e:
                signals.error.emit(str(e))

        threading.Thread(target=worker, daemon=True).start()


    # ==================================================================
    # 3. BƯỚC 1: XỬ LÝ CHÍNH
    # ==================================================================
    def handle_start_processing(self, config_screen) -> None:
        """
        Bắt đầu xử lý (Nối dài hoặc Bổ sung) trong background thread.
        """
        selected_keys = config_screen.get_selected_keys()

        if self._merge_mode == 2:
            if not selected_keys:
                QMessageBox.warning(self.window, "Chưa chọn khóa", "Vui lòng chọn 1 khóa đối chiếu.")
                return
            self._run_left_join_with_mapping(selected_keys, rename_map=getattr(self, "_rename_map", {}))
            return

        # ── Mode 1: Nối dài ──────────────────────────────────────────────
        blocking_keys = selected_keys
        self._sort_config = config_screen.get_sort_config()

        unconfirmed = [
            os.path.basename(p)
            for p in self._selected_files
            if p not in self._file_structures
        ]
        if unconfirmed:
            QMessageBox.warning(
                self.window,
                "Chưa xác nhận cấu trúc",
                "Các file sau chưa xác nhận cấu trúc:\n" + "\n".join(unconfirmed),
            )
            return

        # Lưu mapping vào learning manager
        for struct in self._file_structures.values():
            for orig, std in struct.get("column_mapping", {}).items():
                if std and std != "Bỏ qua (Ignore)":
                    self.lm.remember_mapping(orig, std)

        config_screen.set_scan_button_state(False)
        config_screen.update_progress("⏳ Đang xử lý, vui lòng chờ...", 0)

        confirmed_structs = list(self._file_structures.values())

        # --- PyQt6: Tạo Signal object để giao tiếp giữa Thread và Main UI ---
        class ScanSignals(QObject):
            progress = pyqtSignal(float)
            finished = pyqtSignal(object, list)
            memory_error = pyqtSignal()
            generic_error = pyqtSignal(str)

        signals = ScanSignals()

        # Kết nối tín hiệu (Connect)
        signals.progress.connect(lambda p: config_screen.update_progress("", p * 100))
        signals.finished.connect(lambda df, pairs: self._on_scan_complete(df, pairs, config_screen))

        def on_memory_error():
            QMessageBox.critical(
                self.window,
                "Hết bộ nhớ RAM",
                "Máy tính không đủ RAM để xử lý lượng dữ liệu này.\n"
                "Hãy thử:\n"
                "• Chia nhỏ file đầu vào\n"
                "• Đóng bớt ứng dụng khác\n"
                "• Nâng cấp RAM",
            )
            config_screen.set_scan_button_state(True)
            config_screen.update_progress("❌ Hết RAM – vui lòng chia nhỏ dữ liệu.", 0)

        def on_generic_error(e):
            QMessageBox.critical(self.window, "Lỗi xử lý", f"Chi tiết lỗi:\n{e}")
            config_screen.set_scan_button_state(True)
            config_screen.update_progress("❌ Đã xảy ra lỗi.", 0)

        signals.memory_error.connect(on_memory_error)
        signals.generic_error.connect(on_generic_error)

        def worker() -> None:
            try:
                df, pairs = self.processor.merge_and_deduplicate(
                    file_paths=self._selected_files,
                    final_mapping={},
                    selected_standards=self.STANDARD_COLUMNS,
                    blocking_keys=blocking_keys,
                    confirmed_structures=confirmed_structs,
                    # Gọi emit để gửi tín hiệu cập nhật UI
                    progress_callback=lambda p: signals.progress.emit(p),
                )
                signals.finished.emit(df, pairs)
            except MemoryError:
                signals.memory_error.emit()
            except BaseException as exc:
                import traceback
                err_msg = f"{type(exc).__name__}: {exc}\n\n{traceback.format_exc()[-800:]}"
                signals.generic_error.emit(err_msg)

        threading.Thread(target=worker, daemon=True).start()

    def _on_scan_complete(
        self,
        df: pd.DataFrame,
        pairs: List[Tuple[int, int]],
        dedup_screen,
    ) -> None:
        """Callback về main thread sau khi quét xong."""
        # Loại bỏ cặp đối xứng / tự so
        seen: set = set()
        clean: List[Tuple[int, int]] = []
        for a, b in pairs:
            pk = (min(a, b), max(a, b))
            if pk not in seen and a != b:
                seen.add(pk)
                clean.append((a, b))

        self._raw_combined_df = df
        self._suspect_pairs = clean
        self._current_pair_index = 0
        self._rows_to_delete.clear()
        self._merge_decisions.clear()

        total_rows = len(df)
        n_pairs = len(clean)

        if not clean:
            dedup_screen.update_progress(
                f"✅ Hoàn tất – {total_rows} dòng, không phát hiện trùng lặp.", 100
            )
            QMessageBox.information(
                self.window,
                "Kết quả quét",
                f"✅ Đã quét {total_rows} bản ghi.\nKhông tìm thấy cặp nào trùng nhau.",
            )
            self._goto_final()
        else:
            # Auto-merge
            from services.data_processor import DataProcessor as DP
            merged_df, n_deleted = DP.auto_merge_pairs(
                df=df,
                suspect_pairs=clean,
                standard_columns=self.STANDARD_COLUMNS,
            )
            self._raw_combined_df = merged_df
            self._rows_to_delete = set()
            self._merge_decisions = []
            self._suspect_pairs = []
            dedup_screen.update_progress(
                f"✅ Tự động gộp {n_deleted} bản ghi trùng – còn lại {len(merged_df)} dòng.", 100
            )
            QMessageBox.information(
                self.window,
                "Kết quả quét",
                f"✅ Đã tự động gộp {n_deleted} bản ghi trùng.\n"
                f"Kết quả cuối: {len(merged_df)} bản ghi.",
            )
            self._goto_final()

    # ==================================================================
    # 4. BƯỚC 2: DUYỆT TỪNG CẶP (nếu cần — hiện đang auto-merge)
    # ==================================================================
    def handle_approve(self, review_screen) -> None:
        if self._current_pair_index >= len(self._suspect_pairs):
            return
        id1, id2 = self._suspect_pairs[self._current_pair_index]
        self._merge_decisions.append((id1, id2))
        self._rows_to_delete.add(id2)
        self._current_pair_index += 1
        self._load_pair_to_review(review_screen)

    def handle_reject(self, review_screen) -> None:
        self._current_pair_index += 1
        self._load_pair_to_review(review_screen)

    def handle_skip_all(self, review_screen) -> None:
        reply = QMessageBox.question(
            self.window,
            "Bỏ qua tất cả",
            f"Bỏ qua {len(self._suspect_pairs) - self._current_pair_index} cặp còn lại?\n"
            "Tất cả sẽ được coi là 2 người khác nhau.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._current_pair_index = len(self._suspect_pairs)
            self._load_pair_to_review(review_screen)

    def _load_pair_to_review(self, review_screen) -> None:
        # Bỏ qua cặp đã có bên trong rows_to_delete
        while self._current_pair_index < len(self._suspect_pairs):
            id1, id2 = self._suspect_pairs[self._current_pair_index]
            if id1 not in self._rows_to_delete and id2 not in self._rows_to_delete:
                break
            self._current_pair_index += 1

        if self._current_pair_index >= len(self._suspect_pairs):
            merged = len(self._rows_to_delete)
            QMessageBox.information(
                self.window,
                "Hoàn tất duyệt",
                f"✅ Đã duyệt toàn bộ {len(self._suspect_pairs)} cặp.\n"
                f"Đã hợp nhất {merged} bản ghi trùng lặp.",
            )
            self._goto_final()
            return

        df = self._raw_combined_df
        id1, id2 = self._suspect_pairs[self._current_pair_index]
        row1 = df[df["row_id"] == id1].iloc[0].to_dict()
        row2 = df[df["row_id"] == id2].iloc[0].to_dict()
        review_screen.load_pair(
            row1_data=row1,
            row2_data=row2,
            pair_index=self._current_pair_index,
            total_pairs=len(self._suspect_pairs),
            merged_count=len(self._rows_to_delete),
        )

    # ==================================================================
    # 5. KẾT QUẢ & XUẤT FILE
    # ==================================================================
    def _build_final_df(self) -> pd.DataFrame:
        """Xây dựng DataFrame sạch cuối cùng."""
        if self._raw_combined_df is None:
            return pd.DataFrame(columns=self.STANDARD_COLUMNS)
        from services.data_processor import DataProcessor as DP
        df = DP.apply_approved_merges(
            df=self._raw_combined_df,
            rows_to_delete=self._rows_to_delete,
            standard_columns=self.STANDARD_COLUMNS,
            merge_decisions=self._merge_decisions,
        )
        if "row_id" in df.columns:
            df = df.drop(columns=["row_id"])
        # Sắp xếp theo cấu hình người dùng
        sort_col = self._sort_config.get("col")
        if sort_col and sort_col in df.columns:
            df = df.sort_values(by=sort_col, ascending=self._sort_config.get("ascending", True),
                                na_position="last", ignore_index=True)
        return df.reset_index(drop=True)

    def handle_export(self) -> None:
        """Xuất file kết quả người."""
        if self._raw_combined_df is None:
            QMessageBox.warning(self.window, "Chưa có dữ liệu", "Không có dữ liệu để xuất.")
            return
        try:
            final_df = self._build_final_df()
            self._export_file(final_df)
        except Exception as exc:
            QMessageBox.critical(self.window, "Lỗi lưu file", f"Không thể lưu kết quả:\n{exc}")

    def _export_file(self, clean_df: pd.DataFrame) -> None:
        save_path, _ = QFileDialog.getSaveFileName(
            self.window,
            "Lưu file Excel kết quả",
            "ket_qua_tong_hop.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not save_path:
            return
        try:
            self._export_formatted_excel(clean_df, save_path)
            QMessageBox.information(
                self.window,
                "Xuất thành công",
                f"✅ Đã lưu {len(clean_df)} bản ghi sạch tại:\n{save_path}",
            )
        except Exception as exc:
            QMessageBox.critical(self.window, "Lỗi xuất file", f"Không thể ghi file:\n{exc}")

    def _export_formatted_excel(self, df: pd.DataFrame, save_path: str) -> None:
        """Xuất Excel có định dạng đẹp (giữ nguyên logic từ main_gui.py cũ)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import math

        STD_COLS = set(self.STANDARD_COLUMNS)
        C_HEADER_STD  = "1B3A6B"
        C_HEADER_MISC = "3C5A8A"
        C_STRIPE_ODD  = "F4F8FC"
        C_STRIPE_EVEN = "FFFFFF"
        C_GREEN_BG    = "E6F4EA"
        C_RED_BG      = "FDEDED"
        C_GREEN_FG    = "1B5E20"
        C_RED_FG      = "B71C1C"
        C_WHITE       = "FFFFFF"

        def make_fill(hex_c):
            return PatternFill("solid", fgColor=hex_c)

        def thin_border():
            s = Side(style="thin", color="D0D7E3")
            return Border(left=s, right=s, top=s, bottom=s)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kết quả tổng hợp"

        cols = list(df.columns)
        n_cols = len(cols)

        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        title_cell = ws["A1"]
        title_cell.value = "KẾT QUẢ TỔNG HỢP DỮ LIỆU"
        title_cell.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
        title_cell.fill = make_fill(C_HEADER_STD)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for ci, col_name in enumerate(cols, 1):
            is_std = col_name in STD_COLS
            cell = ws.cell(row=2, column=ci, value=col_name)
            cell.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
            cell.fill = make_fill(C_HEADER_STD if is_std else C_HEADER_MISC)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[2].height = 22

        STATUS_COLS = {"Tình trạng", "Kết quả kiểm tra", "Kết quả"}
        CENTER_COLS = {"STT", "Ngày Sinh", "Ngày nộp", "Số CCCD/ID", "Tình trạng", "Kết quả kiểm tra", "Kết quả"}

        for row_i, row_data in enumerate(df.itertuples(index=False), start=3):
            stripe = C_STRIPE_ODD if row_i % 2 == 1 else C_STRIPE_EVEN
            for ci, col_name in enumerate(cols, 1):
                val = row_data[ci - 1]
                if val is None or (isinstance(val, float) and math.isnan(val)):
                    display = ""
                else:
                    display = val
                cell = ws.cell(row=row_i, column=ci, value=display)
                cell.font = Font(name="Arial", size=10)
                cell.border = thin_border()
                val_str = str(display).strip().lower()
                if col_name in STATUS_COLS and val_str:
                    if any(k in val_str for k in ("đủ", "đạt", "pass", "x")):
                        cell.fill = make_fill(C_GREEN_BG)
                        cell.font = Font(name="Arial", size=10, color=C_GREEN_FG, bold=True)
                    elif any(k in val_str for k in ("không", "trượt", "fail")):
                        cell.fill = make_fill(C_RED_BG)
                        cell.font = Font(name="Arial", size=10, color=C_RED_FG, bold=True)
                    else:
                        cell.fill = make_fill(stripe)
                else:
                    cell.fill = make_fill(stripe)
                cell.alignment = Alignment(
                    horizontal="center" if col_name in CENTER_COLS else "left",
                    vertical="center",
                    indent=0 if col_name in CENTER_COLS else 1,
                )
            ws.row_dimensions[row_i].height = 16

        for ci, col_name in enumerate(cols, 1):
            col_letter = get_column_letter(ci)
            sample = df[col_name].head(200).astype(str).str.len()
            max_len = max(len(col_name), int(sample.max()) if not sample.empty else 0)
            ws.column_dimensions[col_letter].width = max(8, min(40, max_len + 2)) * 1.15

        ws.freeze_panes = "A3"

        ws2 = wb.create_sheet("Thống kê")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 14
        ws2.merge_cells("A1:B1")
        t = ws2["A1"]
        t.value = "THỐNG KÊ TỔNG HỢP"
        t.font = Font(name="Arial", bold=True, size=12, color=C_WHITE)
        t.fill = make_fill(C_HEADER_STD)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 24

        stats = [
            ("Tổng số bản ghi", len(df)),
            ("Số cột dữ liệu", n_cols),
            ("Cột chuẩn",    sum(1 for c in cols if c in STD_COLS)),
            ("Cột bổ sung",  sum(1 for c in cols if c not in STD_COLS)),
        ]
        for ri, (label, val) in enumerate(stats, start=2):
            c1 = ws2.cell(row=ri, column=1, value=label)
            c2 = ws2.cell(row=ri, column=2, value=val)
            bg = C_STRIPE_ODD if ri % 2 == 1 else C_STRIPE_EVEN
            for c in (c1, c2):
                c.font = Font(name="Arial", size=10)
                c.fill = make_fill(bg)
                c.border = thin_border()
                c.alignment = Alignment(
                    horizontal="left" if c == c1 else "center",
                    vertical="center", indent=1 if c == c1 else 0,
                )
            ws2.row_dimensions[ri].height = 16

        wb.save(save_path)

    # ==================================================================
    # 6. CHẾ ĐỘ 2 – LEFT JOIN
    # ==================================================================
    def handle_drag_drop_confirmed(self, screen) -> None:
        """Thực hiện Sequential Left Join trực tiếp từ màn hình kéo thả."""
        rename_map = screen.get_column_mapping()
        key_columns = screen.get_key_columns()

        if not key_columns:
            QMessageBox.warning(self.window, "Chưa có khóa", "Vui lòng bấm nút 🔑 để chọn ít nhất 1 Khóa đối chiếu.")
            return

        # Snapshot file_structures tại thời điểm này để truyền vào worker
        file_structures_snapshot = dict(self._file_structures)

        self.window.update_status(message="⏳  Đang gộp chọn lọc...", progress=5)

        class JoinSignals(QObject):
            progress = pyqtSignal(str, float)
            finished = pyqtSignal(object)
            error = pyqtSignal(str)

        signals = JoinSignals()
        signals.progress.connect(lambda m, p: self.window.update_status(message=m, progress=p))
        signals.finished.connect(lambda df: self.app_ctrl.show_left_join_result(df, key_columns))
        signals.error.connect(lambda err: QMessageBox.critical(self.window, "Lỗi xử lý", f"Gộp chọn lọc thất bại:\n{err}"))

        def progress_cb(msg: str, pct: float) -> None:
            signals.progress.emit(msg, pct)

        def worker() -> None:
            try:
                if not rename_map:
                    result_df = self.processor.run_sequential_left_join(
                        file_paths=self._selected_files,
                        key_columns=key_columns,
                        file_structures=file_structures_snapshot,
                        progress_callback=progress_cb,
                    )
                else:
                    result_df = self._run_rename_join(
                        key_columns, rename_map, progress_cb,
                        file_structures=file_structures_snapshot,
                    )

                self._left_join_result_df = result_df
                signals.finished.emit(result_df)
            except Exception as e:
                signals.error.emit(str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _run_rename_join(self, key_columns, rename_map, progress_cb, file_structures: dict = None):
        """Sequential join khi cột khóa trong file phụ có tên khác File Gốc.
        Sử dụng load_clean_flat_data() để đảm bảo column mapping và date normalization.
        """
        from services.data_processor import DataProcessor as _DP

        def _coalesce_duplicates(df):
            if not df.columns.has_duplicates:
                return df
            import numpy as np
            new_df = pd.DataFrame(index=df.index)
            for col in df.columns.unique():
                subset = df[col]
                if isinstance(subset, pd.DataFrame):
                    subset = subset.replace(r'^\s*$', np.nan, regex=True)
                    new_df[col] = subset.bfill(axis=1).iloc[:, 0]
                else:
                    new_df[col] = subset
            return new_df

        def _read_clean(fp):
            """Dùng load_clean_flat_data nếu có structure đầy đủ, fallback pd.read_excel."""
            struct = (file_structures or {}).get(fp, {})
            if struct and struct.get("file_path"):
                df = self.processor.load_clean_flat_data(struct)
            else:
                sheet = struct.get("sheet_name") or 0
                hrow = struct.get("header_row", 0)
                try:
                    df = pd.read_excel(fp, sheet_name=sheet, header=hrow, dtype=str)
                except Exception:
                    df = pd.read_excel(fp, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            df = _coalesce_duplicates(df)
            return df

        file_paths = self._selected_files
        master_df = _read_clean(file_paths[0])

        aux_dfs = []
        for fp in file_paths[1:]:
            try:
                df_aux = _read_clean(fp)
                
                # Loại bỏ các cột hiện có trùng với đích đổi tên để tránh tạo cột trùng lặp 
                # CHỈ KHI cột nguồn (source) thực sự tồn tại trong file.
                active_renames = {src: tgt for src, tgt in rename_map.items() if src in df_aux.columns}
                targets = set(active_renames.values())
                cols_to_drop = [c for c in df_aux.columns if c in targets and c not in active_renames]
                df_aux = df_aux.drop(columns=cols_to_drop)
                
                df_aux = df_aux.rename(columns=active_renames)
                # Đảm bảo không còn cột nào trùng tên (coalesce thay vì xóa mất dữ liệu)
                df_aux = _coalesce_duplicates(df_aux)
                
                aux_dfs.append((fp, df_aux))
            except Exception as e:
                progress_cb(f"⚠️  Bỏ qua {os.path.basename(fp)}: {e}", 10)

        JOIN_KEY = "__join_key__"

        def _make_key(df, keys):
            """Tạo chuỗi khóa đối chiếu.
            - Cột ngày sinh: chuẩn hóa qua _normalize_date (DD/MM/YYYY).
            - Cột khác: strip + lower.
            - Hàng có key trống → UUID toàn cục duy nhất để tránh ghép nhầm.
            """
            import uuid as _uuid
            parts = []
            for k in keys:
                if k not in df.columns:
                    # Cột khóa không tồn tại → chuỗi rỗng
                    parts.append(pd.Series([""] * len(df), index=df.index))
                    continue
                col_data = df[k]
                if isinstance(col_data, pd.DataFrame):
                    col_data = col_data.iloc[:, 0]
                col_vals = col_data.fillna("").astype(str)
                k_norm = k.lower().strip()
                if any(w in k_norm for w in ("ngay sinh", "ngaysinh", "sinh", "dob", "birthday", "date")):
                    # Chuẩn hóa ngày sinh về DD/MM/YYYY để so khớp chính xác
                    col_vals = col_vals.apply(_DP._normalize_date)
                else:
                    col_vals = col_vals.apply(_DP._clean_text)
                parts.append(col_vals)
            if not parts:
                return pd.Series(
                    [str(_uuid.uuid4()) for _ in range(len(df))],
                    index=df.index,
                )
            raw = parts[0].str.cat(parts[1:], sep="||") if len(parts) > 1 else parts[0]
            # Hàng có key trống → gán UUID duy nhất để không bao giờ khớp với bất kỳ file nào khác
            empty_mask = raw.str.strip() == ""
            result = raw.copy().reset_index(drop=True)
            empty_mask = empty_mask.reset_index(drop=True)
            result.loc[empty_mask] = [str(_uuid.uuid4()) for _ in range(int(empty_mask.sum()))]
            return result

        master_df[JOIN_KEY] = _make_key(master_df, key_columns)
        progress_cb("✅  File Gốc đã nạp. Đang gộp file phụ...", 15)

        n_aux = max(1, len(aux_dfs))
        for i, (fp, df_aux) in enumerate(aux_dfs, start=1):
            pct = 15 + int(i / n_aux * 75)
            progress_cb(f"🔗  Đang gộp File {i + 1}...", pct)
            df_aux[JOIN_KEY] = _make_key(df_aux, key_columns)

            # --- Kiểm tra tỉ lệ khớp khóa ---
            master_keys = set(master_df[JOIN_KEY].dropna().astype(str))
            aux_keys = set(df_aux[JOIN_KEY].dropna().astype(str))
            
            import re
            uuid_pattern = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)
            master_valid_keys = {k for k in master_keys if not uuid_pattern.match(k)}
            aux_valid_keys = {k for k in aux_keys if not uuid_pattern.match(k)}
            
            matched = master_valid_keys.intersection(aux_valid_keys)
            if not matched and master_valid_keys and aux_valid_keys:
                raise ValueError(
                    f"File '{os.path.basename(fp)}' không có dòng nào khớp khóa với File Gốc!\n\n"
                    "Nguyên nhân thường do chọn sai Cột khóa (ví dụ: cột 'Vị trí' ở File Gốc là phòng ban, "
                    "nhưng ở File Phụ lại là chức danh).\n\n"
                    "Vui lòng Quay lại và chỉ chọn 'Họ và Tên' làm khóa đối chiếu, bỏ chọn các cột không đồng nhất."
                )


            # ── Fill-priority cho cột trùng tên ────────────────────────
            data_cols = [c for c in df_aux.columns if c not in key_columns and c != JOIN_KEY]
            existing = set(master_df.columns)
            overlap_cols = [c for c in data_cols if c in existing]
            new_cols = [c for c in data_cols if c not in existing]

            slim = df_aux[[JOIN_KEY] + data_cols].drop_duplicates(subset=[JOIN_KEY], keep="first")

            if new_cols:
                slim_new = slim[[JOIN_KEY] + new_cols]
                master_df = master_df.merge(slim_new, on=JOIN_KEY, how="left",
                                            suffixes=("", f"_dup_{i + 1}"))

            if overlap_cols:
                tmp_names = {c: f"__aux_{c}__" for c in overlap_cols}
                slim_overlap = slim[[JOIN_KEY] + overlap_cols].rename(columns=tmp_names)
                master_df = master_df.merge(slim_overlap, on=JOIN_KEY, how="left")
                for c in overlap_cols:
                    tmp = f"__aux_{c}__"
                    if tmp not in master_df.columns:
                        continue
                    master_vals = master_df[c].fillna("").astype(str).str.strip()
                    aux_vals = master_df[tmp].fillna("").astype(str).str.strip()
                    empty_mask = master_vals.str.lower().isin({"", "nan", "none", "nat"})
                    master_df[c] = master_vals.where(~empty_mask, other=aux_vals)
                    master_df.drop(columns=[tmp], inplace=True)

        if JOIN_KEY in master_df.columns:
            master_df = master_df.drop(columns=[JOIN_KEY])
        master_df = master_df.where(master_df.notna(), other=None).reset_index(drop=True)
        progress_cb(f"🎉  Hoàn tất! {len(master_df)} dòng, {len(master_df.columns)} cột.", 100)
        return master_df

    def handle_export_join(self) -> None:
        """Xuất kết quả Left Join."""
        if self._left_join_result_df is None or self._left_join_result_df.empty:
            QMessageBox.warning(self.window, "Không có dữ liệu", "Chưa có dữ liệu để xuất.")
            return
        save_path, _ = QFileDialog.getSaveFileName(
            self.window,
            "Lưu kết quả Gộp Chọn Lọc",
            "ket_qua_gop_chon_loc.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not save_path:
            return
        try:
            export_df = self._left_join_result_df.copy()
            export_df = export_df.where(export_df.notna(), other=None)
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="Kết quả")
            QMessageBox.information(self.window, "Xuất thành công", f"✅  Đã lưu {len(export_df)} dòng tại:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self.window, "Lỗi xuất file", str(e))

    def handle_back_drag_drop_mapping(self) -> None:
        """
        Từ LeftJoinResultScreen -> quay về màn hình DragDropMappingScreen.
        Tái sử dụng cột đã có trong self._file_structures (không đọc lại file).
        """
        self.window.update_status("Quay lại màn hình Kéo-thả ghép cột...", progress=50)
        self._goto_drag_drop_mapping()
        self.window.update_status("", progress=0)

    # ==================================================================
    # 7. NAVIGATION HELPERS (gọi AppController)
    # ==================================================================
    def _goto_final(self) -> None:
        if self._raw_combined_df is None:
            return
        total_input = len(self._raw_combined_df)
        n_deleted = len(self._rows_to_delete)
        final_df = self._build_final_df()

        # Ghi log
        self.lm.log_session(
            files_processed=len(self._selected_files),
            rows_input=total_input,
            rows_output=len(final_df),
            pairs_found=len(self._suspect_pairs),
            pairs_merged=n_deleted,
        )

        if self.app_ctrl:
            self.app_ctrl.show_people_final(
                final_df=final_df,
                total_input=total_input,
                n_deleted=n_deleted,
                n_files=len(self._selected_files),
            )