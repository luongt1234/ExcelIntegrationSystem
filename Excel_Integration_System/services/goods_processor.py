"""
goods_processor.py
==================
Xử lý logic gộp hàng hóa thông minh:
  • Đọc file Excel đầu vào và file danh mục chuẩn
  • Tính độ tương đồng rapidfuzz
  • Quy tắc: >95% → APPROVED auto, 60-94% → PENDING, <60% → không gợi ý
  • Áp dụng mapping đã duyệt và xuất file Excel sạch
"""

from __future__ import annotations

import os
from typing import Dict, List, Optional, Tuple

import pandas as pd

try:
    from rapidfuzz import fuzz, process as rfprocess
    _HAS_RAPIDFUZZ = True
except ImportError:
    _HAS_RAPIDFUZZ = False

from .goods_learning import GoodsLearningManager, STATUS_APPROVED, STATUS_PENDING, STATUS_REJECTED

THRESHOLD_AUTO   = 95.0   # >  này → tự APPROVED
THRESHOLD_REVIEW = 60.0   # >= này → đưa vào PENDING
# < THRESHOLD_REVIEW → không gợi ý / giữ nguyên


class GoodsProcessor:
    """Xử lý gộp hàng hóa và xuất Excel."""

    def __init__(self, learning_manager: GoodsLearningManager):
        self.lm = learning_manager

    # ------------------------------------------------------------------
    # ĐỌC FILE
    # ------------------------------------------------------------------
    def read_input_file(self, file_path: str, name_column: str,
                        sheet_name: Optional[str] = None,
                        header_row: int = 0) -> pd.DataFrame:
        """Đọc file Excel đầu vào, trả về DataFrame đầy đủ."""
        kwargs: dict = {"header": header_row}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        df = pd.read_excel(file_path, **kwargs)
        df.columns = [str(c).strip() for c in df.columns]
        if name_column not in df.columns:
            raise ValueError(f"Cột '{name_column}' không tồn tại trong file. Các cột có sẵn: {list(df.columns)}")
        return df

    def read_catalog_file(self, file_path: str, name_column: str,
                          sheet_name: Optional[str] = None,
                          header_row: int = 0) -> List[str]:
        """Đọc danh mục chuẩn, trả về danh sách tên."""
        kwargs: dict = {"header": header_row}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        df = pd.read_excel(file_path, **kwargs)
        df.columns = [str(c).strip() for c in df.columns]
        if name_column not in df.columns:
            raise ValueError(f"Cột '{name_column}' không tồn tại trong file danh mục.")
        return df[name_column].dropna().astype(str).str.strip().tolist()

    def get_sheet_names(self, file_path: str) -> List[str]:
        """Lấy danh sách sheet names."""
        try:
            xl = pd.ExcelFile(file_path)
            return xl.sheet_names
        except Exception:
            return ["Sheet1"]

    def get_column_names(self, file_path: str, sheet_name: Optional[str] = None,
                         header_row: int = 0) -> List[str]:
        """Lấy danh sách cột trong file."""
        kwargs: dict = {"header": header_row, "nrows": 0}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        df = pd.read_excel(file_path, **kwargs)
        return [str(c).strip() for c in df.columns]

    # ------------------------------------------------------------------
    # TÍNH ĐỘ TƯƠNG ĐỒNG
    # ------------------------------------------------------------------
    def _similarity(self, a: str, b: str) -> float:
        """Tính độ tương đồng giữa 2 chuỗi (0-100)."""
        if not _HAS_RAPIDFUZZ:
            # Fallback đơn giản nếu không có rapidfuzz
            a_l, b_l = a.lower().strip(), b.lower().strip()
            if a_l == b_l:
                return 100.0
            common = sum(1 for c in set(a_l) if c in b_l)
            return min(common / max(len(a_l), len(b_l), 1) * 100, 99.0)
        return fuzz.token_sort_ratio(a, b)

    def suggest_matches(
        self,
        input_names: List[str],
        catalog_names: List[str],
        progress_callback=None,
    ) -> List[Dict]:
        """
        Với mỗi tên đầu vào, tìm tên chuẩn gần nhất trong danh mục.
        Trả về list dict: {raw_name, best_match, similarity, status}
        Cũng ghi kết quả vào GoodsLearningManager.
        """
        results = []
        total = len(input_names)
        catalog_clean = [c.strip() for c in catalog_names if c.strip()]

        for i, raw in enumerate(input_names):
            raw = str(raw).strip()
            if not raw or raw.lower() in ("nan", "none", ""):
                continue

            if progress_callback:
                progress_callback(i / total)

            # Kiểm tra đã có trong DB chưa
            existing = self.lm.get_status_for_item(raw)
            if existing:
                results.append({
                    "raw_name":     existing["raw_name"],
                    "best_match":   existing["standard_name"],
                    "similarity":   existing["similarity"],
                    "status":       existing["status"],
                })
                continue

            # Tìm khớp tốt nhất
            if not catalog_clean:
                self.lm.set_status(raw, "", STATUS_PENDING, 0.0)
                results.append({"raw_name": raw, "best_match": "", "similarity": 0.0, "status": STATUS_PENDING})
                continue

            if _HAS_RAPIDFUZZ:
                match_result = rfprocess.extractOne(
                    raw, catalog_clean,
                    scorer=fuzz.token_sort_ratio,
                    score_cutoff=0,
                )
                best_name  = match_result[0] if match_result else ""
                best_score = float(match_result[1]) if match_result else 0.0
            else:
                scores = [(c, self._similarity(raw, c)) for c in catalog_clean]
                best_name, best_score = max(scores, key=lambda x: x[1], default=("", 0.0))

            # Phân loại
            if best_score >= THRESHOLD_AUTO:
                status = STATUS_APPROVED
            elif best_score >= THRESHOLD_REVIEW:
                status = STATUS_PENDING
            else:
                status = STATUS_PENDING   # Điểm thấp → vẫn pending, người dùng tự quyết
                best_name = ""            # Không gợi ý

            self.lm.set_status(raw, best_name, status, best_score)
            results.append({
                "raw_name":   raw,
                "best_match": best_name,
                "similarity": best_score,
                "status":     status,
            })

        if progress_callback:
            progress_callback(1.0)

        return results

    # ------------------------------------------------------------------
    # ÁP DỤNG MAPPING VÀO DATAFRAME
    # ------------------------------------------------------------------
    def apply_approved_mappings(self, df: pd.DataFrame, name_column: str,
                                approved_map: Dict[str, str],
                                output_column: str = "Tên chuẩn") -> pd.DataFrame:
        """
        Thêm cột 'Tên chuẩn' vào DataFrame dựa trên approved_map.
        Những tên nằm trong REJECTED hoặc chưa duyệt → giữ nguyên tên gốc.
        """
        df = df.copy()
        rejected = {item["raw_name"] for item in self.lm.get_rejected_items()}

        def _map(raw):
            raw_s = str(raw).strip()
            if raw_s in rejected:
                return raw_s          # Không gộp
            return approved_map.get(raw_s, raw_s)

        df[output_column] = df[name_column].apply(_map)
        return df

    # ------------------------------------------------------------------
    # XUẤT EXCEL ĐẸP
    # ------------------------------------------------------------------
    def export_clean_excel(self, df: pd.DataFrame, save_path: str,
                           name_column: str = "", output_column: str = "Tên chuẩn") -> None:
        """Xuất DataFrame ra Excel có định dạng."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback: xuất không format
            df.to_excel(save_path, index=False)
            return

        C_HEADER  = "1B3A6B"
        C_STD_COL = "1E88E5"   # nền tiêu đề cột Tên chuẩn
        C_ODD     = "F4F8FC"
        C_EVEN    = "FFFFFF"
        C_WHITE   = "FFFFFF"

        def fill(hex_c):
            return PatternFill("solid", fgColor=hex_c)

        def border():
            s = Side(style="thin", color="D0D7E3")
            return Border(left=s, right=s, top=s, bottom=s)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kết quả hàng hóa"

        cols = list(df.columns)
        n_cols = len(cols)

        # Tiêu đề
        ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
        t = ws["A1"]
        t.value = "KẾT QUẢ GỘP HÀNG HÓA THÔNG MINH"
        t.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
        t.fill = fill(C_HEADER)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Header cột
        for ci, col_name in enumerate(cols, 1):
            is_std = col_name == output_column
            cell = ws.cell(row=2, column=ci, value=col_name)
            cell.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
            cell.fill = fill(C_STD_COL if is_std else C_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border()
        ws.row_dimensions[2].height = 22

        # Dữ liệu
        for row_i, row_data in enumerate(df.itertuples(index=False), start=3):
            stripe = C_ODD if row_i % 2 == 1 else C_EVEN
            for ci, col_name in enumerate(cols, 1):
                val = row_data[ci - 1]
                if val is None or (isinstance(val, float) and __import__("math").isnan(val)):
                    val = ""
                cell = ws.cell(row=row_i, column=ci, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = fill(stripe)
                cell.border = border()
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row_i].height = 16

        # Auto-fit cột
        for ci, col_name in enumerate(cols, 1):
            sample = df[col_name].head(200).astype(str).str.len()
            max_len = max(len(col_name), int(sample.max()) if not sample.empty else 0)
            ws.column_dimensions[get_column_letter(ci)].width = max(10, min(45, max_len + 2)) * 1.15

        ws.freeze_panes = "A3"
        wb.save(save_path)