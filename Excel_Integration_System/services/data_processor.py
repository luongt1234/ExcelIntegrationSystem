import re
import copy
import os
import logging
import pandas as pd
from typing import List, Dict, Tuple, Optional, Callable
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from rapidfuzz import fuzz
from .learning_manager import LocalLearningManager

try:
    import openpyxl  # type: ignore
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

try:
    from unidecode import unidecode as _unidecode
    def _norm(text: str) -> str:
        """Chuẩn hóa: bỏ dấu, bỏ khoảng trắng thừa, về chữ thường."""
        return re.sub(r'\s+', ' ', _unidecode(str(text))).strip().lower()
except ImportError:
    import unicodedata
    def _norm(text: str) -> str:
        nfkd = unicodedata.normalize('NFKD', str(text))
        no_accent = ''.join(c for c in nfkd if not unicodedata.combining(c))
        return re.sub(r'\s+', ' ', no_accent).strip().lower()

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Kiểu dữ liệu cho audit event (backward-compatible: chỉ thêm, không thay thế)
# ---------------------------------------------------------------------------
AuditEvent = Dict  # {keep_id, drop_id, column, action, old_value, new_value}


class DataProcessor:
    """
    Bộ xử lý trung tâm v3:
    - Zero-Loss Data Retention: giữ TOÀN BỘ cột từ mọi file đầu vào.
    - Smart Aggregation: FILL_NA khi keep trống, CONFLICT_RESOLVED khi
      cả hai có giá trị khác nhau (nối bằng '|').
    - Tương thích ngược hoàn toàn với main_gui.py hiện tại.
    """

    FUZZY_THRESHOLD = 92.0
    MAPPING_THRESHOLD = 0.35
    SCAN_SAFE_LIMIT = 5000

    # Dấu phân cách khi nối giá trị xung đột
    CONFLICT_SEPARATOR = " | "

    # -----------------------------------------------------------------------
    # Từ khóa dùng để CHẤM ĐIỂM dòng header (bản không dấu, chữ thường).
    # Dùng _norm() khi so khớp để hỗ trợ mọi biến thể Unicode.
    # -----------------------------------------------------------------------
    HEADER_DETECT_KEYWORDS: List[str] = [
        # STT / số thứ tự
        "stt",
        # Họ tên – mọi biến thể không dấu
        "ho va ten", "ho ten", "ho va ten ung vien", "ten ung vien",
        "ho_ten", "hoten", "fullname", "full name", "ho va ten day du",
        # Ngày sinh
        "ngay sinh", "nam sinh", "sinh ngay", "dob", "date of birth",
        # CCCD / CMND
        "cccd", "cmnd", "can cuoc", "can cuoc cong dan", "so dinh danh",
        "so cccd",
        # Điện thoại
        "so dien thoai", "dien thoai", "sdt", "phone", "mobile", "tel",
        # Email
        "email", "thu dien tu", "e-mail", "mail",
        # Địa chỉ
        "dia chi", "noi o", "thuong tru", "address",
        # Vị trí / đơn vị
        "vi tri", "chuc danh", "position", "job", "don vi", "phong ban",
        # Trình độ / bằng cấp
        "trinh do", "bang cap", "van bang", "degree", "education", "chuyen mon",
        # Ghi chú
        "ghi chu", "note", "notes", "remarks",
        # Kết quả / tình trạng
        "ket qua", "tinh trang", "trang thai", "status",
    ]

    # Từ khóa header ứng viên (dùng cho suggest mapping)
    HEADER_KEYWORDS: Dict[str, List[str]] = {
        "Họ và Tên": [
            "họ tên", "họ và tên", "họ và tên ứng viên", "tên ứng viên",
            "họ và tên đầy đủ", "người dự tuyển", "fullname", "full name",
            "ho va ten", "ho ten", "ho va ten ung vien", "hoten", "ten ung vien",
            "ho va ten day du", "ten", "name", "họ và tên người nộp",
            "thí sinh", "ứng viên", "ung vien",
        ],
        "Ngày Sinh": [
            "ngày sinh", "năm sinh", "sinh ngày", "dob", "birthday",
            "date of birth", "ngay sinh", "nam sinh", "sinh ngay",
            "ngày sinh (dd/mm/yyyy)", "ngày sinh dd/mm/yyyy",
            "năm sinh (yyyy)", "ngày tháng năm sinh",
        ],
        "Số CCCD/ID": [
            "cccd", "cmnd", "căn cước", "căn cước công dân", "số định danh",
            "id", "số cccd", "số cmnd", "số căn cước", "so cccd", "so cmnd",
            "can cuoc cong dan", "so dinh danh", "chứng minh nhân dân",
            "chứng minh thư", "cmnd/cccd", "cccd/cmnd",
        ],
        "Số Điện Thoại": [
            "số điện thoại", "điện thoại", "sđt", "phone", "mobile", "tel",
            "so dien thoai", "dien thoai", "sdt", "số dt", "số đt",
            "điện thoại liên lạc", "sđt liên lạc", "số điện thoại liên lạc",
            "phone number", "di động", "di dong", "hotline",
        ],
        "Email": [
            "email", "thư điện tử", "e-mail", "mail",
            "địa chỉ email", "email liên hệ", "email liên lạc",
            "dia chi email", "thu dien tu",
        ],
        "Địa Chỉ": [
            "địa chỉ", "nơi ở", "thường trú", "liên hệ", "address",
            "dia chi", "noi o", "thuong tru", "địa chỉ hiện tại",
            "địa chỉ thường trú", "nơi cư trú", "tỉnh/thành phố",
            "tinh thanh pho", "quê quán", "que quan",
        ],
        "Vị trí": [
            "vị trí", "vị trí ứng tuyển", "chức danh", "công việc",
            "position", "job", "vi tri", "vi tri ung tuyen", "chuc danh",
            "chức vụ", "chuc vu", "vị trí dự tuyển", "ngành nghề",
            "job title", "role",
        ],
        "Đơn vị": [
            "đơn vị", "phòng ban", "bộ phận", "department",
            "don vi", "phong ban", "bo phan", "cơ quan", "co quan",
            "trường", "trung tâm", "công ty", "chi nhánh",
            "đơn vị công tác", "nơi công tác",
        ],
        "Ngày nộp": [
            "ngày nộp", "ngày ứng tuyển", "ngày nhận hồ sơ", "apply date",
            "ngay nop", "ngay ung tuyen", "ngay nhan ho so",
            "ngày đăng ký", "ngày đăng ký dự tuyển", "thời gian nộp",
            "submission date", "applied date",
        ],
        "Tình trạng": [
            "tình trạng", "trạng thái", "kết quả", "status",
            "tinh trang", "trang thai", "ket qua",
            "kết quả kiểm tra", "kết quả xét tuyển", "đạt/không đạt",
            "pass/fail", "đủ điều kiện", "kết luận",
        ],
        "Văn bằng": [
            "văn bằng", "bằng cấp", "trình độ", "chuyên môn",
            "degree", "education", "van bang", "bang cap", "trinh do",
            "trình độ học vấn", "trình độ chuyên môn", "bằng tốt nghiệp",
            "cấp độ học vấn", "bằng đại học", "bằng thạc sỹ",
            "highest education", "qualification",
        ],
        "Ghi Chú": [
            "ghi chú", "ghi chu", "note", "notes", "remarks",
            "chú thích", "chu thich", "ghi chú thêm", "thông tin khác",
            "thongong tin khac", "bổ sung", "comment", "comments",
        ],
    }

    # Các cột quan trọng không được nối xung đột âm thầm
    CRITICAL_COLUMNS = {"Số CCCD/ID", "Ngày Sinh", "Email", "Số Điện Thoại"}

    @staticmethod
    def _coalesce_duplicated_columns(df: "pd.DataFrame") -> "pd.DataFrame":
        """Gộp các cột bị trùng tên bằng cách lấy giá trị không null đầu tiên (bfill theo hàng)."""
        if not df.columns.duplicated().any():
            return df
        new_cols = {}
        for c in df.columns.unique():
            col_data = df[c]
            if isinstance(col_data, pd.DataFrame):
                # Replace whitespace-only strings and common null-like strings with pd.NA to ensure bfill works
                temp = col_data.replace(r'^\s*$', pd.NA, regex=True)
                temp = temp.replace(["nan", "NaN", "None", "NaT", "nat"], pd.NA)
                new_cols[c] = temp.bfill(axis=1).iloc[:, 0]
            else:
                new_cols[c] = col_data
        return pd.DataFrame(new_cols)

    def __init__(self, learning_manager: LocalLearningManager):
        self.lm = learning_manager

    # ==================================================================
    # PHẦN 1b: QUÉT CẤU TRÚC FILE & GỢI Ý MAPPING (Confirm-first flow)
    # ==================================================================

    # -----------------------------------------------------------------------
    # Từ khóa dùng để nhận biết dòng KHÔNG phải ứng viên (lọc bỏ)
    # -----------------------------------------------------------------------
    JUNK_ROW_KEYWORDS: List[str] = [
        "trinh do dao tao", "van bang tot nghiep", "bang diem",
        "truc tiep", "vi tri chuyen vien", "vi tri", "phu luc",
        "bang kiem tra", "trinh do dao tao", "van bang",
        "ket qua kiem tra", "dieu kien tieu chuan",
        "so tt", "stt",
    ]

    def scan_excel_structure(
        self,
        file_path: str,
        sheet_name=None,
        max_rows: int = 20,
    ) -> dict:
        """
        Quét nhanh cấu trúc file Excel và trả về gợi ý để UI hiển thị.

        Returns dict:
            file_path, sheet_names, selected_sheet,
            suggested_header_row, confidence_score,
            header_start_row, header_end_row, data_start_row,
            preview_rows (list[list]), detected_columns (list[str]),
            suggested_mapping (dict[str, str])
        """
        result = {
            "file_path": file_path,
            "sheet_names": [],
            "selected_sheet": None,
            "suggested_header_row": 0,
            "confidence_score": 0.0,
            "header_start_row": 0,
            "header_end_row": 0,
            "data_start_row": 1,
            "preview_rows": [],
            "detected_columns": [],
            "suggested_mapping": {},
            "header_warning": "",
        }
        # Ưu tiên dùng pandas ExcelFile (nhanh, ít RAM hơn openpyxl full-load)
        # để lấy sheet names. Chỉ fallback sang openpyxl khi pandas thất bại.
        try:
            with pd.ExcelFile(file_path) as xf:
                sheet_names = xf.sheet_names
        except Exception:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            except Exception as e:
                raise ValueError(f"Không thể đọc file. Hãy đảm bảo file đang không bị mở bởi phần mềm khác (như Excel). Chi tiết: {e}")

        result["sheet_names"] = sheet_names

        # Chọn sheet: ưu tiên sheet được chỉ định, rồi đến sheet có nhiều dữ liệu nhất
        if sheet_name and sheet_name in sheet_names:
            chosen_sheet = sheet_name
        else:
            chosen_sheet = sheet_names[0]
            best_count = -1
            for sn in sheet_names:
                try:
                    tmp = pd.read_excel(file_path, sheet_name=sn, header=None, nrows=5)
                    cnt = tmp.notna().sum().sum()
                    if cnt > best_count:
                        best_count = cnt
                        chosen_sheet = sn
                except Exception:
                    pass

        result["selected_sheet"] = chosen_sheet

        # Đọc preview
        try:
            preview_df = pd.read_excel(
                file_path, sheet_name=chosen_sheet, header=None, nrows=max_rows
            )
        except Exception:
            return result

        result["preview_rows"] = preview_df.fillna("").astype(str).values.tolist()

        # Dò header (có thể nhiều tầng)
        header_info = self.detect_header_range_from_preview(
            preview_df, sheet_name=chosen_sheet
        )
        header_start = header_info["header_start_row"]
        header_end = header_info["header_end_row"]
        data_start = header_info["data_start_row"]
        confidence = header_info["confidence"]

        result["suggested_header_row"] = header_start
        result["confidence_score"] = confidence
        result["header_start_row"] = header_start
        result["header_end_row"] = header_end
        result["data_start_row"] = data_start

        # Cảnh báo UI khi không tìm được header tin cậy
        if confidence == 0.0:
            result["header_warning"] = (
                "⚠️ Không tự nhận diện được dòng tiêu đề, "
                "vui lòng chọn Header Row thủ công."
            )
        else:
            result["header_warning"] = ""

        # Flatten header nhiều tầng thành danh sách tên cột
        flat_cols = self.flatten_header_rows_from_file(
            file_path, chosen_sheet, header_start, header_end
        )
        if not flat_cols:
            flat_cols = self.flatten_header_rows(preview_df, header_start, header_end)
        flat_cols = self._dedupe_column_names(flat_cols)
        header_vals = [c for c in flat_cols if c and c.lower() not in ("", "nan", "nat", "none")]

        result["detected_columns"] = header_vals

        logger.debug(
            "[scan_excel_structure] sheet=%r header_start=%d header_end=%d "
            "data_start=%d confidence=%.2f detected_columns=%s",
            chosen_sheet, header_start, header_end, data_start, confidence, header_vals,
        )

        # Gợi ý mapping
        result["suggested_mapping"] = self.suggest_column_mapping(
            header_vals, list(self.HEADER_KEYWORDS.keys())
        )

        return copy.deepcopy(result)

    # ------------------------------------------------------------------
    # Nhận diện vùng header (có thể nhiều tầng) và dòng bắt đầu dữ liệu
    # ------------------------------------------------------------------

    def detect_header_range_from_preview(
        self, preview_df: "pd.DataFrame", sheet_name: str = ""
    ) -> dict:
        """
        Trả về dict:
            header_start_row, header_end_row, data_start_row, confidence
        Hỗ trợ header nhiều tầng: tìm dòng bắt đầu header, rồi quét tiếp
        các dòng liền kề có thể là header phụ (ít ô hơn nhưng không phải dữ liệu ứng viên).
        """
        kw_set = set(self.HEADER_DETECT_KEYWORDS)

        best_row = 0
        best_score = -1
        best_hits = 0

        for i in range(len(preview_df)):
            row = preview_df.iloc[i]
            vals_norm = [
                _norm(v)
                for v in row
                if _norm(str(v)) not in ("", "nan", "nat", "none")
            ]
            if not vals_norm:
                continue

            kw_hits = sum(1 for v in vals_norm if v in kw_set)
            score = kw_hits * 10

            if score > best_score:
                best_score = score
                best_row = int(i)
                best_hits = kw_hits

        if best_hits < 2:  # Yêu cầu tối thiểu 2 keyword để tránh nhầm dòng tiêu đề báo cáo
            logger.warning(
                "[header_detect] sheet=%r → không đủ keyword (hits=%d < 2), mặc định dùng dòng 0.",
                sheet_name or "?", best_hits,
            )
            return {
                "header_start_row": 0,
                "header_end_row": 0,
                "data_start_row": 1,
                "confidence": 0.0,
            }

        # Confidence
        row_vals = [
            v for v in preview_df.iloc[best_row]
            if _norm(str(v)) not in ("", "nan", "nat", "none")
        ]
        confidence = round(min(1.0, best_hits / max(1, len(row_vals))), 2)

        # Quét các dòng tiếp theo để xác định header_end_row
        # Một dòng vẫn là header phụ nếu:
        #   - Không phải dữ liệu ứng viên (không có tên người rõ ràng)
        #   - Có chứa từ khóa header hoặc phần lớn ô là text mô tả (không phải ngày/số)
        header_end = best_row
        max_probe = min(best_row + 8, len(preview_df) - 1)

        for j in range(best_row + 1, max_probe + 1):
            row_j = preview_df.iloc[j]
            vals_j = [str(v).strip() for v in row_j if str(v).strip() not in ("", "nan", "NaT", "None")]
            if not vals_j:
                # Dòng trống → kết thúc vùng header
                break
            # Kiểm tra có phải dòng dữ liệu ứng viên không
            if self._row_is_candidate_data(row_j):
                break
            # Kiểm tra dòng có phải tiêu đề phụ không
            if self._row_is_sub_header(row_j):
                header_end = j
            else:
                # Không rõ → dừng mở rộng header
                break

        data_start = header_end + 1

        logger.debug(
            "[header_detect] sheet=%r → header_start=%d header_end=%d data_start=%d confidence=%.2f",
            sheet_name or "?", best_row, header_end, data_start, confidence,
        )

        return {
            "header_start_row": best_row,
            "header_end_row": header_end,
            "data_start_row": data_start,
            "confidence": confidence,
        }

    def _row_is_candidate_data(self, row: "pd.Series") -> bool:
        """
        Dòng là dữ liệu ứng viên thật khi:
          - Có một STT là số nguyên nhỏ (1..9999) ở 3 ô đầu, VÀ
          - Có ít nhất một ô trông như NGÀY (date) hoặc tên người (>=2 từ chữ)
            không chứa các cụm tiêu đề phụ.
        Điều này tránh nhầm các dòng header phụ kiểu
        "Trực tiếp / Bưu điện / Bằng Đại học / Bằng Thạc sỹ" thành dữ liệu.
        """
        sub_header_kws = (
            "vi tri", "phu luc", "bang kiem tra", "trinh do dao tao",
            "van bang", "ket qua kiem tra", "dieu kien", "truc tiep",
            "buu dien", "bang dai hoc", "bang thac sy", "bang tien sy",
            "chung chi", "kinh nghiem", "du dieu kien", "khong du dieu kien",
            "doi tuong uu tien", "cc cntt", "ccnn",
        )
        cells = list(row)
        # 1) Có STT số nguyên ở 3 ô đầu?
        has_stt = False
        for v in cells[:3]:
            s = str(v).strip()
            if s in ("", "nan", "NaT", "None"):
                continue
            try:
                n = float(s)
                if 0 < n < 10000 and abs(n - int(n)) < 1e-9:
                    has_stt = True
                    break
            except (ValueError, TypeError):
                pass
        if not has_stt:
            return False
        # 2) Có ngày hoặc tên người thật?
        for v in cells:
            s = str(v).strip()
            if s in ("", "nan", "NaT", "None"):
                continue
            # Ngày: chuỗi có / hoặc - và parse được
            if ("/" in s or "-" in s) and self._looks_like_date_or_number(s):
                return True
            # Tên người Vietnamese-like: 2-6 từ chữ cái, không chứa kw header phụ
            words = s.split()
            if 2 <= len(words) <= 6:
                v_norm = _norm(s)
                if not any(kw in v_norm for kw in sub_header_kws):
                    # Phải có chữ cái thuần (không phải toàn số/ký hiệu)
                    if any(ch.isalpha() for ch in s):
                        return True
        return False

    def _row_is_sub_header(self, row: "pd.Series") -> bool:
        """Kiểm tra dòng có phải là header phụ (tiêu đề bổ sung) không."""
        kw_set = set(self.HEADER_DETECT_KEYWORDS)
        junk_kw = set(self.JUNK_ROW_KEYWORDS)
        sub_header_substrs = (
            "trinh do dao tao", "van bang tot nghiep", "bang diem",
            "ket qua kiem tra", "dieu kien tieu chuan",
            "truc tiep", "buu dien", "vi tri ung tuyen",
            "bang dai hoc", "bang thac sy", "bang tien sy",
            "chung chi", "cc cntt", "ccnn",
            "du dieu kien", "khong du dieu kien", "doi tuong uu tien",
            "kinh nghiem", "van bang", "giay to co lien quan",
            "ngay sinh", "sdt", "dia chi", "don ung tuyen",
            "ho va ten", "ngay nop", "tinh day du",
        )
        vals_norm = [
            _norm(str(v))
            for v in row
            if _norm(str(v)) not in ("", "nan", "nat", "none")
        ]
        if not vals_norm:
            return False
        kw_hits = 0
        for v in vals_norm:
            if v in kw_set or v in junk_kw:
                kw_hits += 1
                continue
            if any(s in v for s in sub_header_substrs):
                kw_hits += 1
        if kw_hits > 0:
            return True
        all_short_text = all(
            len(v.split()) <= 6 and not self._looks_like_date_or_number(v)
            for v in vals_norm
        )
        return all_short_text and len(vals_norm) >= 2

    @staticmethod
    def _looks_like_date_or_number(val: str) -> bool:
        """Kiểm tra chuỗi có dạng ngày hoặc số không."""
        import re as _re
        val = val.strip()
        if _re.match(r'^\d+([.,/\-]\d+)*$', val):
            return True
        if pd.notna(pd.to_datetime(val, errors='coerce', dayfirst=True)):
            return True
        return False

    # ------------------------------------------------------------------
    # Flatten header nhiều tầng thành danh sách tên cột
    # ------------------------------------------------------------------

    @staticmethod
    def flatten_header_rows(
        df: "pd.DataFrame",
        header_start: int,
        header_end: int,
    ) -> List[str]:
        """
        Fallback flatten dựa trên DataFrame (không có thông tin merged cells).
        Quy tắc:
        - KHÔNG ffill ngang (sẽ làm lan tên cha sang cột không liên quan).
        - Với mỗi cột: nối các giá trị không rỗng theo chiều dọc bằng " - ".
        - Bỏ phần trùng lặp.
        Dùng `flatten_header_rows_from_file` khi có file để bù merged cells.
        """
        if header_start > header_end:
            header_end = header_start
        header_block = df.iloc[header_start: header_end + 1].copy()
        num_cols = header_block.shape[1]
        result_cols = []
        for col_idx in range(num_cols):
            parts: List[str] = []
            for row_idx in range(len(header_block)):
                cell = str(header_block.iloc[row_idx, col_idx]).strip()
                if cell.lower() in ("", "nan", "nat", "none"):
                    continue
                if cell not in parts:
                    parts.append(cell)
            result_cols.append(" - ".join(parts) if parts else f"Col_{col_idx}")
        return result_cols

    # ------------------------------------------------------------------
    # Flatten header sử dụng openpyxl để bù merged cells một cách chính xác
    # ------------------------------------------------------------------
    @staticmethod
    def flatten_header_rows_from_file(
        file_path: str,
        sheet_name: Optional[str],
        header_start: int,
        header_end: int,
    ) -> List[str]:
        """
        Đọc trực tiếp các ô header bằng openpyxl, mở rộng merged cells
        (top-left value fill vào cả vùng merge) rồi nối dọc theo từng cột.
        Đây là cách CHÍNH XÁC để xử lý header nhiều tầng có merged cells,
        tránh việc ffill ngang làm lan giá trị cha sang cột không liên quan.
        """
        if not _HAS_OPENPYXL:
            return []
        if header_start > header_end:
            header_end = header_start
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            max_col = ws.max_column or 0
            n_rows = header_end - header_start + 1
            # Khởi tạo ma trận ô (raw)
            matrix: List[List] = []
            for r0 in range(header_start, header_end + 1):
                r = r0 + 1  # openpyxl 1-indexed
                row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
                matrix.append(row_vals)
            # Áp merged ranges: với mỗi range có giao với band header,
            # gán giá trị top-left vào tất cả ô trong giao đó.
            for mr in ws.merged_cells.ranges:
                top_val = ws.cell(mr.min_row, mr.min_col).value
                for r in range(mr.min_row, mr.max_row + 1):
                    r0 = r - 1
                    if not (header_start <= r0 <= header_end):
                        continue
                    band_idx = r0 - header_start
                    for c in range(mr.min_col, min(mr.max_col, max_col) + 1):
                        c0 = c - 1
                        if 0 <= c0 < max_col:
                            matrix[band_idx][c0] = top_val
            wb.close()

            cols: List[str] = []
            for c in range(max_col):
                parts: List[str] = []
                for r in range(n_rows):
                    v = matrix[r][c]
                    if v is None:
                        continue
                    s = re.sub(r"\s+", " ", str(v)).strip()
                    if not s or s.lower() in ("nan", "nat", "none"):
                        continue
                    if s not in parts:
                        parts.append(s)
                cols.append(" - ".join(parts) if parts else f"Col_{c}")
            return cols
        except Exception as exc:
            logger.warning("[flatten_header_rows_from_file] lỗi: %s", exc)
            return []

    @staticmethod
    def _dedupe_column_names(cols: List[str]) -> List[str]:
        """Đảm bảo tên cột là duy nhất bằng cách thêm hậu tố _2, _3..."""
        seen: Dict[str, int] = {}
        out: List[str] = []
        for c in cols:
            base = c
            if base not in seen:
                seen[base] = 1
                out.append(base)
            else:
                seen[base] += 1
                out.append(f"{base}_{seen[base]}")
        return out

    def detect_header_row_from_preview(
        self, preview_df: "pd.DataFrame", sheet_name: str = ""
    ) -> "Tuple[int, float]":
        """Backward-compat wrapper: trả về (header_start_row, confidence)."""
        info = self.detect_header_range_from_preview(preview_df, sheet_name=sheet_name)
        return info["header_start_row"], info["confidence"]


    def suggest_column_mapping(
        self,
        header_values: List[str],
        target_standards: List[str],
    ) -> Dict[str, str]:
        """
        Gợi ý mapping từ header thực tế sang schema chuẩn.
        Ưu tiên: (1) memory SQLite, (2) exact match, (3) alias match
                 (bản gốc + bản không dấu + từng token),
                 (4) partial-contains match, (5) TF-IDF + Fuzzy kết hợp.

        Thay đổi so với phiên bản cũ:
        - Trọng số: fuzz*0.55 + tfidf*0.45 (fuzzy tốt hơn với tiếng Việt ngắn).
        - Thêm bước 4: kiểm tra alias có nằm trong tên cột dài không (partial).
        - Thêm partial-fuzzy riêng từng token của tên cột flatten nhiều tầng.
        - MAPPING_THRESHOLD vẫn 0.35 (đã tốt), thêm fast-accept >= 0.70.
        """
        result: Dict[str, str] = {}

        # Xây alias_map: alias_str → standard_name
        alias_map: Dict[str, str] = {}
        for std, aliases in self.HEADER_KEYWORDS.items():
            for alias in aliases:
                alias_map[alias.lower()] = std
                alias_map[_norm(alias)] = std

        for hv in header_values:
            hv_clean = re.sub(r'\s+', ' ', str(hv)).strip()
            hv_lower = hv_clean.lower()
            hv_norm = _norm(hv_clean)

            # ── 1. Memory (SQLite learned mappings) ──────────────────────
            learned = self.lm.get_learned_mapping(hv_clean)
            if learned and learned in target_standards:
                result[hv] = learned
                continue

            # ── 2. Exact match (case-insensitive) ────────────────────────
            exact = next(
                (ts for ts in target_standards if hv_lower == ts.strip().lower()),
                None,
            )
            if exact:
                result[hv] = exact
                continue

            # ── 3. Alias exact match (gốc + không dấu) ───────────────────
            alias_hit = alias_map.get(hv_lower) or alias_map.get(hv_norm)
            if alias_hit and alias_hit in target_standards:
                result[hv] = alias_hit
                continue

            # ── 3b. Token match: "Khác - SĐT", "Điều kiện - Văn bằng" ───
            # Ưu tiên token CUỐI (cụ thể nhất sau flatten nhiều tầng)
            tokens = [t.strip() for t in re.split(r"\s*[-/]\s*", hv_clean) if t.strip()]
            token_hit = None
            for tok in reversed(tokens):
                t_norm = _norm(tok)
                t_low = tok.lower()
                hit = alias_map.get(t_low) or alias_map.get(t_norm)
                if hit and hit in target_standards:
                    token_hit = hit
                    break
            if token_hit:
                result[hv] = token_hit
                continue

            # ── 3c. Partial-contains: alias nằm trong tên cột dài ────────
            # Ví dụ: cột "Ngày sinh (dd/mm/yyyy)" chứa alias "ngày sinh"
            partial_hit = None
            for alias_key, std_name in alias_map.items():
                if std_name not in target_standards:
                    continue
                if len(alias_key) >= 4 and (
                    alias_key in hv_lower or alias_key in hv_norm
                ):
                    partial_hit = std_name
                    break
            if partial_hit:
                result[hv] = partial_hit
                continue

            # ── 4. TF-IDF + Fuzzy (fallback) ─────────────────────────────
            # Với header flatten nhiều tầng, lấy token score tốt nhất thay vì
            # dùng toàn bộ chuỗi dài (tránh cosine bị kéo thấp bởi padding).
            all_probe_strings = [hv_norm] + [_norm(t) for t in tokens if _norm(t)]

            best_target: Optional[str] = None
            best_score: float = 0.0

            for ts in target_standards:
                ts_norm = _norm(ts)
                best_probe_score = 0.0

                for probe in all_probe_strings:
                    fuzz_score = fuzz.token_sort_ratio(probe, ts_norm) / 100.0
                    try:
                        vec = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4))
                        mat = vec.fit_transform([probe, ts_norm])
                        tfidf_score = float(
                            cosine_similarity(mat[0:1], mat[1:2])[0][0]
                        )
                    except Exception:
                        tfidf_score = 0.0
                    # Trọng số mới: fuzzy nặng hơn (tốt hơn với tiếng Việt ngắn)
                    combined = fuzz_score * 0.55 + tfidf_score * 0.45
                    if combined > best_probe_score:
                        best_probe_score = combined

                if best_probe_score > best_score:
                    best_score = best_probe_score
                    best_target = ts

                # Fast-accept: điểm cao rõ ràng → không cần so tiếp
                if best_score >= 0.70:
                    break

            result[hv] = (
                best_target
                if (best_score >= self.MAPPING_THRESHOLD and best_target)
                else "Bỏ qua (Ignore)"
            )

        return result

    def load_clean_flat_data(self, confirmed_structure: dict) -> "pd.DataFrame":
        """
        Đọc và làm sạch dữ liệu Excel dựa trên cấu trúc đã xác nhận từ người dùng.
        Hỗ trợ header nhiều tầng: dùng header_start_row, header_end_row, data_start_row
        nếu có trong confirmed_structure.

        confirmed_structure keys:
            file_path, sheet_name,
            header_row (int) – dòng header bắt đầu (0-based),
            header_start_row (int, optional) – alias của header_row,
            header_end_row (int, optional) – dòng header kết thúc,
            data_start_row (int, optional) – dòng dữ liệu thực bắt đầu,
            selected_columns (list[str]),   # tên đã flatten (gốc)
            column_mapping (dict[str, str]) # tên gốc → tên chuẩn
        """
        path = confirmed_structure["file_path"]
        sheet = confirmed_structure.get("sheet_name")
        mapping = confirmed_structure.get("column_mapping", {})
        selected = confirmed_structure.get("selected_columns", [])

        # Xác định vùng header và data
        header_start = int(
            confirmed_structure.get("header_start_row",
            confirmed_structure.get("header_row", 0))
        )
        header_end = int(
            confirmed_structure.get("header_end_row", header_start)
        )
        data_start = int(
            confirmed_structure.get("data_start_row", header_end + 1)
        )

        is_multilevel = header_end > header_start
        kwargs = {} if not sheet else {"sheet_name": sheet}

        if is_multilevel:
            # --- Đọc toàn bộ file không set header ---
            raw_df = pd.read_excel(path, header=None, **kwargs)

            # Flatten header block
            flat_cols = self.flatten_header_rows_from_file(
                path, sheet, header_start, header_end
            )
            if not flat_cols:
                flat_cols = self.flatten_header_rows(raw_df, header_start, header_end)
            flat_cols = self._dedupe_column_names(flat_cols)

            # Lấy dữ liệu từ data_start trở xuống
            data_df = raw_df.iloc[data_start:].copy().reset_index(drop=True)
            # Đặt tên cột
            n_cols = min(len(flat_cols), data_df.shape[1])
            col_names = flat_cols[:n_cols]
            # Nếu file có nhiều cột hơn header, đặt tên generic
            for extra in range(n_cols, data_df.shape[1]):
                col_names.append(f"_Col_{extra}")
            data_df.columns = col_names

            logger.debug(
                "[load_clean_flat_data] multilevel sheet=%r header=%d-%d data_start=%d columns=%s",
                sheet, header_start, header_end, data_start, col_names,
            )
        else:
            # --- Header 1 dòng: đọc thông thường ---
            raw_df = pd.read_excel(path, header=None, **kwargs)
            # Lấy tên cột từ dòng header
            flat_cols = self.flatten_header_rows_from_file(
                path, sheet, header_start, header_start
            )
            if not flat_cols:
                flat_cols = self.flatten_header_rows(raw_df, header_start, header_start)
            flat_cols = self._dedupe_column_names(flat_cols)
            data_df = raw_df.iloc[data_start:].copy().reset_index(drop=True)
            n_cols = min(len(flat_cols), data_df.shape[1])
            col_names = flat_cols[:n_cols]
            for extra in range(n_cols, data_df.shape[1]):
                col_names.append(f"_Col_{extra}")
            data_df.columns = col_names

            logger.debug(
                "[load_clean_flat_data] single-level sheet=%r header=%d columns=%s",
                sheet, header_start, col_names,
            )

        df = data_df

        # --- Chuẩn hóa tên cột ---
        df.columns = [re.sub(r'\s+', ' ', str(c)).strip() for c in df.columns]

        # --- Lọc dòng không phải ứng viên ---
        df = self._filter_non_candidate_rows(df)

        # Chỉ giữ cột được chọn
        selected_norm = [re.sub(r'\s+', ' ', str(c)).strip() for c in selected]
        keep_cols = [c for c in selected_norm if c in df.columns]
        if keep_cols:
            df = df[keep_cols].copy()

        # Đổi tên cột theo mapping
        rename_dict = {}
        for k, v in mapping.items():
            if not v or v == "Bỏ qua (Ignore)":
                continue
            k_norm = re.sub(r'\s+', ' ', str(k)).strip()
            if k_norm in df.columns:
                rename_dict[k_norm] = v
        df = df.rename(columns=rename_dict)

        # Gộp các cột trùng tên (do nhiều cột được map về cùng 1 tên)
        df = DataProcessor._coalesce_duplicated_columns(df)

        # Drop dòng trống hoàn toàn
        df = df.dropna(how="all").reset_index(drop=True)

        return df

    # ------------------------------------------------------------------
    # Lọc dòng không phải ứng viên thật
    # ------------------------------------------------------------------

    def _filter_non_candidate_rows(self, df: "pd.DataFrame") -> "pd.DataFrame":
        """
        Bỏ các dòng không phải ứng viên:
          - Dòng trống hoàn toàn
          - Dòng nhóm có STT là số La Mã (I, II, III, IV, V, ...) trong 3 ô đầu
          - Dòng có cụm tiêu đề phụ (Vị trí, PHỤ LỤC, BẢNG KIỂM TRA, Trình độ
            đào tạo, Văn bằng, Bảng điểm, Kết quả kiểm tra, Trực tiếp/Bưu điện,
            Đủ điều kiện, Bằng Đại học/Thạc sỹ/Tiến sỹ, ...) ở cột tên hoặc
            chiếm đa số các ô.
          - Dòng cột Họ và Tên chứa cụm tiêu đề phụ (Vị trí Chuyên viên,
            Trình độ đào tạo, Văn bằng, PHỤ LỤC, BẢNG KIỂM TRA, ...).
        Lưu ý: KHÔNG tự động bỏ chỉ vì cột Họ và Tên trống – một số file
        thực tế có cột tên trống nhưng vẫn có STT, Ngày sinh, SĐT hợp lệ.
        """
        if df.empty:
            return df

        # Tìm cột tên ứng viên (sau khi đã rename hoặc tên gốc)
        name_col = None
        for c in df.columns:
            c_norm = _norm(str(c))
            if c_norm.startswith(("ho va ten", "ho ten")) or c_norm == "ten":
                name_col = c
                break

        junk_phrases = (
            "vi tri chuyen vien", "phu luc", "bang kiem tra",
            "trinh do dao tao", "van bang tot nghiep", "bang diem",
            "ket qua kiem tra", "dieu kien tieu chuan",
            "truc tiep", "buu dien",
            "bang dai hoc", "bang thac sy", "bang tien sy",
            "chung chi", "du dieu kien", "khong du dieu kien",
            "doi tuong uu tien",
        )
        roman_re = re.compile(r"^(i{1,3}|iv|v|vi{1,3}|ix|x|xi{1,3}|xiv|xv)$")

        def cell_str(v) -> str:
            s = "" if v is None else str(v).strip()
            return "" if s.lower() in ("nan", "nat", "none") else s

        def has_numeric_stt(cells) -> bool:
            for v in cells[:3]:
                s = cell_str(v)
                if not s:
                    continue
                try:
                    n = float(s)
                    if 0 < n < 100000 and abs(n - int(n)) < 1e-9:
                        return True
                except (ValueError, TypeError):
                    pass
            return False

        def is_junk_row(row) -> bool:
            cells = list(row)
            vals = [cell_str(v) for v in cells if cell_str(v) != ""]
            if not vals:
                return True

            # 1) Số La Mã trong 3 ô đầu → dòng nhóm
            for v in cells[:3]:
                s = cell_str(v).lower()
                if s and roman_re.match(s):
                    return True

            # 2) Cột Họ và Tên chứa cụm tiêu đề phụ
            if name_col is not None:
                try:
                    nv = cell_str(row[name_col])
                except Exception:
                    nv = ""
                if nv:
                    nv_norm = _norm(nv)
                    for p in junk_phrases:
                        if p in nv_norm:
                            return True

            # 3) Nếu dòng KHÔNG có STT số nguyên ở đầu thì coi là junk khi
            # các ô của nó toàn là cụm tiêu đề phụ (header phụ lọt vào data).
            if not has_numeric_stt(cells):
                vals_norm = [_norm(v) for v in vals]
                junk_hits = sum(
                    1 for vn in vals_norm
                    if any(p in vn for p in junk_phrases)
                )
                if junk_hits >= 1 and junk_hits / max(1, len(vals_norm)) >= 0.5:
                    return True
                # Không có STT và không có dữ liệu định danh nào → bỏ
                if not any(
                    self._looks_like_date_or_number(v) for v in vals
                ):
                    # toàn text, không số, không ngày → header phụ
                    if all(len(v.split()) <= 8 for v in vals):
                        return True

            return False

        # ── Vectorized filtering (thay thế df.apply(axis=1) chậm) ──────────────
        # Bước 1: Dòng trống hoàn toàn
        mask_all_empty = df.apply(
            lambda col: col.map(cell_str) == ""
        ).all(axis=1)

        # Bước 2: Số La Mã trong 3 cột đầu
        first_3_cols = df.columns[:3].tolist()
        mask_roman = pd.Series(False, index=df.index)
        for c in first_3_cols:
            mask_roman |= df[c].map(lambda v: bool(roman_re.match(cell_str(v).lower())))

        # Bước 3: Cột tên chứa junk phrase
        mask_name_junk = pd.Series(False, index=df.index)
        if name_col is not None:
            def _name_is_junk(v) -> bool:
                nv = cell_str(v)
                if not nv:
                    return False
                nv_norm = _norm(nv)
                return any(p in nv_norm for p in junk_phrases)
            mask_name_junk = df[name_col].map(_name_is_junk)

        # Bước 4: Dòng không có STT số nguyên VÀ toàn cụm junk / text ngắn
        # (chỉ với dòng chưa bị đánh dấu bởi bước 1-3 để tiết kiệm thời gian)
        candidates_mask = ~(mask_all_empty | mask_roman | mask_name_junk)
        if candidates_mask.any():
            remaining = df[candidates_mask]

            def _has_stt(row) -> bool:
                return has_numeric_stt(list(row[first_3_cols]))

            def _no_stt_and_junk(row) -> bool:
                if _has_stt(row):
                    return False
                vals = [cell_str(v) for v in row if cell_str(v)]
                if not vals:
                    return True
                vals_norm = [_norm(v) for v in vals]
                junk_hits = sum(1 for vn in vals_norm if any(p in vn for p in junk_phrases))
                if junk_hits >= 1 and junk_hits / max(1, len(vals_norm)) >= 0.5:
                    return True
                if not any(self._looks_like_date_or_number(v) for v in vals):
                    if all(len(v.split()) <= 8 for v in vals):
                        return True
                return False

            # apply chỉ trên subset còn lại → ít dòng hơn nhiều
            mask_no_stt_junk = remaining.apply(_no_stt_and_junk, axis=1)
            mask_extra = pd.Series(False, index=df.index)
            mask_extra[mask_no_stt_junk[mask_no_stt_junk].index] = True
        else:
            mask_extra = pd.Series(False, index=df.index)

        mask = mask_all_empty | mask_roman | mask_name_junk | mask_extra
        filtered = df[~mask].reset_index(drop=True)

        logger.debug(
            "[filter_non_candidate_rows] trước=%d dòng, sau=%d dòng, loại=%d dòng",
            len(df), len(filtered), mask.sum(),
        )
        return filtered

    # ==================================================================
    # PHẦN 1: TRÍCH XUẤT & ÁNH XẠ TIÊU ĐỀ CỘT  (giữ tương thích cũ)
    # ==================================================================

    @staticmethod
    def _detect_header_row(path: str, sheet_name=None, max_scan: int = 15) -> int:
        """
        Tự động phát hiện hàng tiêu đề trong 15 dòng đầu của mỗi sheet.

        Cải tiến: dùng _norm() (bỏ dấu tiếng Việt, lower, collapse whitespace)
        khi so khớp từ khóa → nhận diện đúng "Họ và tên ứng viên", "Họ và tên "
        (dư cách), các bảng có tiêu đề phụ như "PHỤ LỤC I" ở trên.
        """
        try:
            kw_set = set(DataProcessor.HEADER_DETECT_KEYWORDS)
            kwargs = {} if sheet_name is None else {"sheet_name": sheet_name}

            # Đọc preview 15 dòng không set header
            df_raw = pd.read_excel(path, header=None, nrows=max_scan, **kwargs)

            best_row = 0
            best_score = -1

            for i in range(len(df_raw)):
                row = df_raw.iloc[i]
                vals_norm = [
                    _norm(v)
                    for v in row
                    if _norm(str(v)) not in ("", "nan", "nat", "none")
                ]
                if not vals_norm:
                    continue

                kw_hits = sum(1 for v in vals_norm if v in kw_set)
                score = kw_hits * 10

                if score > best_score:
                    best_score = score
                    best_row = int(i)

            logger.debug(
                "[_detect_header_row] path=%r sheet=%r → header_row=%d score=%d",
                path, sheet_name, best_row, best_score,
            )
            return best_row
        except Exception as exc:
            logger.warning("[_detect_header_row] lỗi: %s", exc)
            return 0

    def extract_headers(self, file_paths: List[str]) -> List[str]:
        """Trích xuất tất cả tiêu đề cột độc nhất từ các file đầu vào."""
        all_headers: set = set()
        for path in file_paths:
            try:
                header_row = DataProcessor._detect_header_row(path)
                df = pd.read_excel(path, header=header_row, nrows=0)
                # Chuẩn hóa khoảng trắng tên cột (giữ dấu để hiển thị)
                cols = [re.sub(r'\s+', ' ', str(c)).strip() for c in df.columns]
                all_headers.update(cols)
            except Exception as e:
                print(f"[WARN] Lỗi đọc tiêu đề từ '{path}': {e}")
        return sorted(list(all_headers))

    def compute_semantic_mapping(
        self,
        source_headers: List[str],
        target_standards: List[str],
    ) -> Dict[str, str]:
        """
        Ánh xạ cột thông minh: SQLite memory → khớp chính xác → TF-IDF + Fuzzy.
        Trả về dict {header_goc: header_chuẩn | "Bỏ qua (Ignore)"}.
        """
        result: Dict[str, str] = {}

        for sh in source_headers:
            sh_clean = sh.strip()

            learned = self.lm.get_learned_mapping(sh_clean)
            if learned and learned in target_standards:
                result[sh] = learned
                continue

            exact = next(
                (ts for ts in target_standards if sh_clean.lower() == ts.strip().lower()),
                None,
            )
            if exact:
                result[sh] = exact
                continue

            best_target: Optional[str] = None
            best_score: float = 0.0

            for ts in target_standards:
                fuzz_score = fuzz.token_sort_ratio(sh_clean.lower(), ts.lower()) / 100.0
                try:
                    vec = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4))
                    mat = vec.fit_transform([sh_clean.lower(), ts.lower()])
                    tfidf_score = float(cosine_similarity(mat[0:1], mat[1:2])[0][0])
                except Exception:
                    tfidf_score = 0.0

                combined = fuzz_score * 0.4 + tfidf_score * 0.6
                if combined > best_score:
                    best_score = combined
                    best_target = ts

            result[sh] = best_target if (best_score >= self.MAPPING_THRESHOLD and best_target) else "Bỏ qua (Ignore)"

        return result

    # ==================================================================
    # PHẦN 2: GỘP FILE & QUÉT CẶP TRÙNG  (Zero-Loss)
    # ==================================================================

    def merge_and_deduplicate(
        self,
        file_paths: List[str],
        final_mapping: Dict[str, str],
        selected_standards: List[str],
        blocking_keys: List[str],
        confirmed_structures: Optional[List[dict]] = None,
        progress_callback: Optional[Callable[[float], None]] = None,
    ) -> Tuple[pd.DataFrame, List[Tuple[int, int]]]:
        """
        Giai đoạn 1 – đồng bộ cấu trúc cột (ZERO-LOSS: giữ toàn bộ cột phụ).
        Giai đoạn 2 – chuẩn hóa sâu cột selected_standards; strip nhẹ cột phụ.
        Giai đoạn 3 – phát hiện CẶP nghi trùng theo blocking_keys.

        Tham số mới (backward-compatible):
            confirmed_structures: danh sách dict cấu trúc đã người dùng xác nhận.
                Nếu có, dùng load_clean_flat_data() thay vì tự detect header.
                Nếu None, fallback về hành vi detect tự động.

        Trả về:
            combined_df  : DataFrame đã có cột 'row_id' nội bộ (0-indexed).
            suspect_pairs: danh sách (row_id_1, row_id_2) nghi trùng.
        """
        master_dfs: List[pd.DataFrame] = []
        total_files = len(file_paths)
        all_renamed_cols: set = set()

        if confirmed_structures:
            # --- Luồng mới: Dynamic Union – chỉ dùng cột người dùng đã tick ---
            struct_map = {s["file_path"]: s for s in confirmed_structures}

            # Pass 1: xây dựng union schema CHỈ từ cột người dùng tick giữ lại.
            # Cột được map về tên chuẩn → dùng tên chuẩn; cột không map → giữ tên gốc.
            # KHÔNG tự sinh cột mặc định từ selected_standards nếu không có trong file nguồn.
            union_ordered: list = []          # giữ thứ tự xuất hiện, không trùng
            union_set: set = set()

            for path in file_paths:
                struct = struct_map.get(path)
                if struct is None:
                    continue
                mapping = struct.get("column_mapping", {})
                selected = struct.get("selected_columns", [])
                for c in selected:
                    c_norm = re.sub(r'\s+', ' ', str(c)).strip()
                    mapped = mapping.get(c_norm) or mapping.get(c, "Bỏ qua (Ignore)")
                    final_name = mapped if (mapped and mapped != "Bỏ qua (Ignore)") else c_norm
                    if final_name not in union_set:
                        union_ordered.append(final_name)
                        union_set.add(final_name)
                    all_renamed_cols.add(final_name)

            union_schema = union_ordered  # thứ tự từ file đầu tiên, union với file sau

            # Pass 2: load dữ liệu từng file, căn chỉnh theo union_schema
            for idx, path in enumerate(file_paths):
                struct = struct_map.get(path)
                if struct is None:
                    continue
                try:
                    df = self.load_clean_flat_data(struct)
                    df = DataProcessor._coalesce_duplicated_columns(df)
                    # Thêm cột còn thiếu trong union_schema (để NaN); không thêm cột ngoài schema
                    missing_cols = [col for col in union_schema if col not in df.columns]
                    if missing_cols:
                        df = pd.concat(
                            [df, pd.DataFrame(None, index=df.index, columns=missing_cols)],
                            axis=1,
                        )
                    # Chỉ giữ các cột trong union_schema (bảo toàn thứ tự) — không .copy() thêm
                    df = df[[c for c in union_schema if c in df.columns]]
                    master_dfs.append(df)
                except Exception as e:
                    print(f"[WARN] Lỗi đọc file '{path}': {e}")
                if progress_callback:
                    progress_callback((idx + 1) / total_files * 0.35)

        else:
            # --- Fallback: detect tự động (giữ tương thích cũ) ---
            for path in file_paths:
                try:
                    header_row = DataProcessor._detect_header_row(path)
                    df_head = pd.read_excel(path, header=header_row, nrows=0)
                    renamed_cols = [
                        final_mapping.get(c, c) if final_mapping.get(c, "Bỏ qua (Ignore)") != "Bỏ qua (Ignore)" else c
                        for c in df_head.columns.astype(str)
                    ]
                    all_renamed_cols.update(renamed_cols)
                except Exception:
                    pass

            extra_cols = sorted(
                [c for c in all_renamed_cols if c not in selected_standards]
            )
            union_schema = selected_standards + extra_cols

            for idx, path in enumerate(file_paths):
                try:
                    header_row = DataProcessor._detect_header_row(path)
                    df = pd.read_excel(path, header=header_row)
                    rename_dict = {
                        k: v
                        for k, v in final_mapping.items()
                        if v != "Bỏ qua (Ignore)" and k in df.columns
                    }
                    df = df.rename(columns=rename_dict)
                    df = DataProcessor._coalesce_duplicated_columns(df)
                    for col in union_schema:
                        if col not in df.columns:
                            df[col] = None
                    remaining = [c for c in df.columns if c not in union_schema]
                    df = df[union_schema + remaining].copy()
                    master_dfs.append(df)
                except Exception as e:
                    print(f"[WARN] Lỗi xử lý file '{path}': {e}")
                if progress_callback:
                    progress_callback((idx + 1) / total_files * 0.35)

        if not master_dfs:
            fallback_cols = union_schema if union_schema else selected_standards
            empty = pd.DataFrame(columns=["row_id"] + fallback_cols)
            return empty, []

        combined_df = pd.concat(master_dfs, ignore_index=True)

        # Loại bỏ hàng rỗng hoàn toàn (do merged-cells / header phụ còn sót lại)
        combined_df = combined_df.dropna(how="all").reset_index(drop=True)
        # Loại bỏ hàng mà TẤT CẢ cột chuẩn đều rỗng/NaN (hàng vô nghĩa)
        std_cols_present = [c for c in selected_standards if c in combined_df.columns]
        if std_cols_present:
            mask_empty = combined_df[std_cols_present].apply(
                lambda col: col.astype(str).str.strip().isin(["", "nan", "NaT", "None"])
            ).all(axis=1)
            combined_df = combined_df[~mask_empty].reset_index(drop=True)

        # Loại bỏ cột trùng tên (giữ cột đầu tiên): tránh combined_df[col]
        # trả về DataFrame thay vì Series → gây lỗi 'DataFrame' has no attribute 'str'
        combined_df = combined_df.loc[:, ~combined_df.columns.duplicated(keep="first")]

        # ---- Giai đoạn 2: chuẩn hóa dữ liệu -------------------------
        for col in combined_df.columns:
            # Guard: bỏ qua nếu cột vẫn bị trùng (không nên xảy ra sau bước trên)
            if isinstance(combined_df[col], pd.DataFrame):
                continue
            if col in selected_standards:
                # Chuẩn hóa sâu chỉ cho cột chuẩn
                if col == "Ngày Sinh":
                    combined_df[col] = combined_df[col].apply(self._normalize_date)
                else:
                    combined_df[col] = (
                        combined_df[col].fillna("").astype(str).str.strip()
                    )
            else:
                # Cột phụ: chỉ strip nhẹ, giữ nguyên kiểu nếu có thể
                combined_df[col] = combined_df[col].apply(
                    lambda v: str(v).strip() if pd.notna(v) and str(v).strip() not in ("", "nan", "NaT") else ""
                )

        # Chuyển chuỗi rỗng thành None và xóa các cột rỗng hoàn toàn
        for col in combined_df.columns:
            combined_df[col] = combined_df[col].apply(lambda v: None if pd.isna(v) or str(v).strip() == "" else v)
            
        cols_to_keep = [c for c in combined_df.columns if combined_df[c].notna().any()]
        combined_df = combined_df[cols_to_keep]

        # Gắn row_id bất biến
        combined_df.insert(0, "row_id", range(len(combined_df)))

        if progress_callback:
            progress_callback(0.40)

        if not blocking_keys:
            return combined_df, []

        # ---- Giai đoạn 3: quét cặp trùng ----------------------------
        suspect_pairs = self._scan_duplicates(
            combined_df, blocking_keys, selected_standards, progress_callback
        )

        if progress_callback:
            progress_callback(1.0)

        return combined_df, suspect_pairs

    # ==================================================================
    # PHẦN 3: LOGIC LOẠI TRỪ TRÙNG LẶP SAU KHI DUYỆT  (Smart Aggregation)
    # ==================================================================

    @staticmethod
    def apply_approved_merges(
        df: pd.DataFrame,
        rows_to_delete: set,
        standard_columns: List[str],
        merge_decisions: Dict[int, Tuple[int, int]],
        return_audit: bool = False,
    ) -> "pd.DataFrame | Tuple[pd.DataFrame, List[AuditEvent]]":
        """
        Áp dụng toàn bộ quyết định "Hợp nhất" đã được phê duyệt.

        Smart Aggregation:
        - keep trống, drop có dữ liệu  → ghi drop sang keep (FILL_NA).
        - cả hai giống nhau             → giữ nguyên (không audit).
        - cả hai khác nhau              → nối "keep | drop" nếu chưa chứa
                                          (CONFLICT_RESOLVED).

        Tham số mới (backward-compatible):
            return_audit (bool, default=False):
                Nếu True, trả về (df_cleaned, list[AuditEvent]).
                Nếu False (mặc định), trả về df_cleaned như cũ.

        AuditEvent schema:
            {
                "keep_id"   : int,
                "drop_id"   : int,
                "column"    : str,
                "action"    : "FILL_NA" | "CONFLICT_RESOLVED",
                "old_value" : str,   # giá trị keep ban đầu
                "new_value" : str,   # giá trị keep sau khi merge
            }
        """
        SEP = DataProcessor.CONFLICT_SEPARATOR
        df = df.copy()
        # Đảm bảo index unique để df.at[] luôn trả về scalar, không trả về Series
        df = df.set_index("row_id")
        if not df.index.is_unique:
            df = df.reset_index()
            df = df.drop_duplicates(subset=["row_id"], keep="first")
            df = df.set_index("row_id")

        audit_events: List[AuditEvent] = []

        for keep_id, drop_id in merge_decisions:
            if keep_id not in df.index or drop_id not in df.index:
                continue

            # Xử lý tất cả cột có trong DataFrame (Zero-Loss: không giới hạn selected_standards)
            for col in df.columns:
                # Ép về scalar an toàn (tránh "truth value of Series is ambiguous")
                v_keep_raw = DataProcessor._to_scalar(df.at[keep_id, col])
                v_drop_raw = DataProcessor._to_scalar(df.at[drop_id, col])

                v_keep = str(v_keep_raw).strip() if not DataProcessor._is_empty_scalar(v_keep_raw) else ""
                v_drop = str(v_drop_raw).strip() if not DataProcessor._is_empty_scalar(v_drop_raw) else ""

                keep_empty = v_keep == "" or v_keep.lower() == "nan"
                drop_empty = v_drop == "" or v_drop.lower() == "nan"

                if keep_empty and not drop_empty:
                    # FILL_NA
                    df.at[keep_id, col] = v_drop
                    audit_events.append({
                        "keep_id": keep_id,
                        "drop_id": drop_id,
                        "column": col,
                        "action": "FILL_NA",
                        "old_value": "",
                        "new_value": v_drop,
                    })
                elif not keep_empty and not drop_empty and v_keep != v_drop:
                    # CONFLICT_RESOLVED: nối nếu chưa chứa sẵn
                    parts = [p.strip() for p in v_keep.split("|")]
                    if v_drop not in parts:
                        new_val = v_keep + SEP + v_drop
                        df.at[keep_id, col] = new_val
                        audit_events.append({
                            "keep_id": keep_id,
                            "drop_id": drop_id,
                            "column": col,
                            "action": "CONFLICT_RESOLVED",
                            "old_value": v_keep,
                            "new_value": new_val,
                        })
                # else: cả hai rỗng hoặc giống nhau → không làm gì

        df = df.drop(index=[rid for rid in rows_to_delete if rid in df.index])
        df = df.reset_index(drop=False)  # drop=False: đưa row_id trở lại thành cột

        if return_audit:
            return df, audit_events
        return df

    # ==================================================================
    # HELPER CÔNG KHAI: build_merge_audit_events
    # ==================================================================

    @staticmethod
    def build_merge_audit_events(
        df_before: pd.DataFrame,
        merge_decisions: Dict[int, Tuple[int, int]],
    ) -> List[AuditEvent]:
        """
        Helper độc lập: tạo danh sách AuditEvent từ df và merge_decisions
        MÀ KHÔNG thực sự sửa df.

        Dùng khi audit_logger.py hoặc main_gui.py cần preview audit
        trước khi commit.

        Trả về list[AuditEvent] với schema:
            keep_id, drop_id, column, action, old_value, new_value
        """
        SEP = DataProcessor.CONFLICT_SEPARATOR
        df = df_before.set_index("row_id") if "row_id" in df_before.columns else df_before
        events: List[AuditEvent] = []

        for keep_id, drop_id in merge_decisions:
            if keep_id not in df.index or drop_id not in df.index:
                continue
            for col in df.columns:
                v_keep_raw = DataProcessor._to_scalar(df.at[keep_id, col])
                v_drop_raw = DataProcessor._to_scalar(df.at[drop_id, col])
                v_keep = str(v_keep_raw).strip() if not DataProcessor._is_empty_scalar(v_keep_raw) else ""
                v_drop = str(v_drop_raw).strip() if not DataProcessor._is_empty_scalar(v_drop_raw) else ""
                keep_empty = v_keep == "" or v_keep.lower() == "nan"
                drop_empty = v_drop == "" or v_drop.lower() == "nan"

                if keep_empty and not drop_empty:
                    events.append({
                        "keep_id": keep_id, "drop_id": drop_id,
                        "column": col, "action": "FILL_NA",
                        "old_value": "", "new_value": v_drop,
                    })
                elif not keep_empty and not drop_empty and v_keep != v_drop:
                    parts = [p.strip() for p in v_keep.split("|")]
                    if v_drop not in parts:
                        events.append({
                            "keep_id": keep_id, "drop_id": drop_id,
                            "column": col, "action": "CONFLICT_RESOLVED",
                            "old_value": v_keep,
                            "new_value": v_keep + SEP + v_drop,
                        })
        return events

    # ==================================================================
    # HÀM NỘI BỘ
    # ==================================================================

    @staticmethod
    def _to_scalar(val):
        """
        Ép giá trị về scalar an toàn.
        df.at[] có thể trả về Series nếu index không unique → gây lỗi
        'truth value of a Series is ambiguous'. Hàm này lấy phần tử đầu tiên
        nếu val là Series, giữ nguyên nếu đã là scalar.
        """
        if isinstance(val, pd.Series):
            return val.iloc[0] if len(val) > 0 else None
        return val

    # ==================================================================
    # HELPER CÔNG KHAI: auto_merge_pairs
    # ==================================================================

    @staticmethod
    def auto_merge_pairs(
        df: pd.DataFrame,
        suspect_pairs: List[Tuple[int, int]],
        standard_columns: List[str],
    ) -> Tuple[pd.DataFrame, int]:
        """
        Tự động gộp TẤT CẢ các cặp đã được _scan_duplicates xác nhận
        (Họ và Tên + Ngày Sinh đã khớp). Áp dụng Smart Aggregation
        (FILL_NA + CONFLICT_RESOLVED) và XÓA bản ghi trùng thứ hai.

        Khử bắc cầu: nếu A~B và B~C thì gom A, B, C cùng một keep_id.
        Trả về (df_đã_gộp_và_loại_trùng, số_dòng_đã_xóa).
        """
        if not suspect_pairs:
            return df.copy(), 0

        # Union-Find để khử bắc cầu
        parent: Dict[int, int] = {}

        def find(x: int) -> int:
            while parent.get(x, x) != x:
                parent[x] = parent.get(parent.get(x, x), parent.get(x, x))
                x = parent[x]
            parent.setdefault(x, x)
            return x

        def union(a: int, b: int) -> None:
            ra, rb = find(a), find(b)
            if ra == rb:
                return
            # Giữ id nhỏ hơn làm keep (thường là dòng xuất hiện trước)
            keep, drop = (ra, rb) if ra < rb else (rb, ra)
            parent[drop] = keep

        for a, b in suspect_pairs:
            union(int(a), int(b))

        # Sinh merge_decisions: với mỗi root, gắn từng thành viên còn lại
        groups: Dict[int, List[int]] = {}
        all_ids = set()
        for a, b in suspect_pairs:
            all_ids.add(int(a)); all_ids.add(int(b))
        for rid in all_ids:
            root = find(rid)
            groups.setdefault(root, []).append(rid)

        merge_decisions: List[Tuple[int, int]] = []
        rows_to_delete: set = set()
        for root, members in groups.items():
            for m in members:
                if m == root:
                    continue
                merge_decisions.append((root, m))
                rows_to_delete.add(m)

        df_out = DataProcessor.apply_approved_merges(
            df=df,
            rows_to_delete=rows_to_delete,
            standard_columns=standard_columns,
            merge_decisions=merge_decisions,
        )
        return df_out, len(rows_to_delete)

    @staticmethod
    def _is_empty_scalar(val) -> bool:
        """Kiểm tra scalar có rỗng/NaN không (an toàn, không gây ambiguous)."""
        if val is None:
            return True
        try:
            if pd.isna(val):
                return True
        except (TypeError, ValueError):
            pass
        return str(val).strip() in ("", "nan", "NaT", "None")

    @staticmethod
    def _normalize_date(val) -> str:
        """Chuẩn hóa mọi định dạng ngày sinh về DD/MM/YYYY."""
        if pd.isna(val) or str(val).strip() in ("", "nan", "NaT"):
            return ""
        try:
            dt = pd.to_datetime(val, errors="coerce", dayfirst=True)
            if pd.notna(dt):
                return dt.strftime("%d/%m/%Y")
        except Exception:
            pass
        return str(val).strip()

    @staticmethod
    def _clean_text(val) -> str:
        """Loại bỏ khoảng trắng thừa và chuyển thường."""
        if not val:
            return ""
        return " ".join(str(val).lower().split()).strip()

    def _scan_duplicates(
        self,
        df: pd.DataFrame,
        blocking_keys: List[str],
        standard_columns: List[str],
        progress_callback: Optional[Callable[[float], None]],
    ) -> List[Tuple[int, int]]:
        """
        Quét toàn bộ cặp (i, j) nghi trùng theo blocking_keys.
        Sử dụng cơ chế Blocking Group thông minh để giảm độ phức tạp từ O(N²) xuống gần O(N).
        """
        records = df.to_dict(orient="records")
        
        # Pre-clean blocking keys to avoid O(N^2) string processing overhead
        for rec in records:
            for key in blocking_keys:
                if key in rec:
                    rec[key] = self._clean_text(rec[key])

        suspect_pairs: List[Tuple[int, int]] = []
        seen: set = set()

        def is_duplicate_pair(r1: dict, r2: dict) -> bool:
            match_count = 0
            for key in blocking_keys:
                v1 = r1.get(key, "")
                v2 = r2.get(key, "")
                if not v1 or not v2:
                    return False
                if key in ("Ngày Sinh", "Số CCCD/ID", "Số Điện Thoại"):
                    if v1 == v2:
                        match_count += 1
                    else:
                        return False
                else:
                    score = fuzz.token_sort_ratio(v1, v2)
                    if score >= self.FUZZY_THRESHOLD:
                        match_count += 1
                    else:
                        return False
            return match_count == len(blocking_keys)

        exact_keys = [k for k in blocking_keys if k in ("Ngày Sinh", "Số CCCD/ID", "Số Điện Thoại")]

        # Giới hạn kích thước bucket: nếu 1 bucket quá lớn (toàn dòng trống key),
        # lấy mẫu ngẫu nhiên thay vì so O(N²). Ngưỡng 300 → max ~45.000 phép tính/bucket.
        MAX_BUCKET_SIZE = 300

        buckets: Dict[tuple, List[int]] = {}
        for idx, rec in enumerate(records):
            b_key = []
            if exact_keys:
                for k in exact_keys:
                    val = rec.get(k, "")
                    # Nếu key rỗng → KHÔNG gom vào bucket chung; dùng sentinel duy nhất
                    # để tránh mega-bucket của dòng trống (gây O(N²) với blank data).
                    b_key.append(val if val else f"__EMPTY_{idx}__")
            else:
                for k in blocking_keys:
                    val = rec.get(k, "")
                    if not val:
                        b_key.append(f"__EMPTY_{idx}__")
                    else:
                        tokens = sorted(val.split())
                        sorted_val = " ".join(tokens)
                        b_key.append(sorted_val[:2])

            t_key = tuple(b_key)
            if t_key not in buckets:
                buckets[t_key] = []
            buckets[t_key].append(idx)

        # Loại bỏ bucket singleton (không thể tạo cặp) và giới hạn bucket quá lớn
        import random as _random
        filtered_buckets: Dict[tuple, List[int]] = {}
        for t_key, indices in buckets.items():
            # Bỏ bucket singleton và bucket của sentinel rỗng (chứa "__EMPTY_")
            if len(indices) < 2:
                continue
            key_str = " ".join(str(k) for k in t_key)
            if "__EMPTY_" in key_str:
                continue
            if len(indices) > MAX_BUCKET_SIZE:
                logger.warning(
                    "[_scan_duplicates] Bucket kích thước %d > MAX_BUCKET_SIZE=%d, "
                    "lấy mẫu ngẫu nhiên để tránh O(N²).",
                    len(indices), MAX_BUCKET_SIZE,
                )
                indices = _random.sample(indices, MAX_BUCKET_SIZE)
            filtered_buckets[t_key] = indices

        total_buckets = max(1, len(filtered_buckets))
        for b_idx, (_, indices) in enumerate(filtered_buckets.items()):
            m = len(indices)
            for ii in range(m):
                for jj in range(ii + 1, m):
                    i, j = indices[ii], indices[jj]
                    pk = (min(i, j), max(i, j))
                    if pk in seen:
                        continue
                    if is_duplicate_pair(records[i], records[j]):
                        seen.add(pk)
                        suspect_pairs.append((records[i]["row_id"], records[j]["row_id"]))
            if progress_callback:
                progress_callback(0.40 + ((b_idx + 1) / total_buckets) * 0.60)

        return suspect_pairs

    def run_sequential_left_join(
        self,
        file_paths: List[str],
        key_columns: List[str],
        file_structures: dict = None,
        progress_callback=None,
    ) -> "pd.DataFrame":
        """
        Thuật toán Sequential Left Join cuốn chiếu.
        1. Nạp File Gốc (file_paths[0]) -> df_master.
        2. Lần lượt left-join với từng file phụ (file_paths[1:]).
        3. Chuẩn hóa key trước khi join: trim, lowercase, chuẩn hóa ngày.
        4. Cột trùng tên: fill-priority (giữ master, lấy aux nếu master trống).
        5. Kết quả cuối: thay NaN -> None.
        """
        def _report(msg, pct):
            if progress_callback:
                try:
                    progress_callback(msg, pct)
                except Exception:
                    pass

        def _read_file(fp):
            struct = (file_structures or {}).get(fp, {})
            if struct and struct.get("file_path"):
                # Dùng load_clean_flat_data: áp dụng column mapping + lọc dòng không hợp lệ
                try:
                    df = self.load_clean_flat_data(struct)
                    df.columns = [str(c).strip() for c in df.columns]
                    df = df.loc[:, ~df.columns.duplicated(keep='first')]
                    return df
                except Exception:
                    pass  # Fallback xuống bên dưới nếu load_clean_flat_data thất bại
            if struct.get("sheet_name") is not None and struct.get("header_row") is not None:
                # Dùng cấu hình đã xác nhận từ StructureConfirmScreen
                sheet = struct["sheet_name"]
                hrow = struct["header_row"]
                try:
                    df = pd.read_excel(fp, sheet_name=sheet, header=hrow, dtype=str)
                except Exception:
                    df = pd.read_excel(fp, dtype=str)
            else:
                # Fallback: scan tự động
                try:
                    scan = self.scan_excel_structure(fp, max_rows=5)
                    header_row = scan.get("suggested_header_row", 0)
                    sheet = scan.get("selected_sheet", 0)
                    df = pd.read_excel(fp, sheet_name=sheet, header=header_row, dtype=str)
                except Exception:
                    df = pd.read_excel(fp, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            df = df.loc[:, ~df.columns.duplicated(keep='first')]
            return df

        def _make_join_key(df, keys):
            """Tao chuoi khoa doi chieu. Hang co key trong -> UUID toan cuc duy nhat
            de tranh ghep nham cac ban ghi trong giua cac file khac nhau."""
            import uuid as _uuid
            parts = []
            for k in keys:
                if k not in df.columns:
                    parts.append(pd.Series([""] * len(df), index=df.index))
                    continue
                col_data = df[k]
                if isinstance(col_data, pd.DataFrame):
                    col_data = col_data.iloc[:, 0]
                col_vals = col_data.fillna("").astype(str)
                k_norm = k.lower().strip()
                if any(w in k_norm for w in ("ngay sinh", "ngaysinh", "sinh", "dob", "birthday", "date")):
                    col_vals = col_vals.apply(self._normalize_date)
                else:
                    col_vals = col_vals.apply(self._clean_text)
                parts.append(col_vals)
            sep = "||"
            raw = parts[0].str.cat(parts[1:], sep=sep) if len(parts) > 1 else parts[0]
            # Hang co key trong -> gan UUID duy nhat de khong bao gio khop voi bat ky file nao khac
            empty_mask = raw.str.strip() == ""
            result = raw.copy().reset_index(drop=True)
            empty_mask = empty_mask.reset_index(drop=True)
            result.loc[empty_mask] = [str(_uuid.uuid4()) for _ in range(int(empty_mask.sum()))]
            return result

        _report("📂  Đang đọc File Gốc...", 5)
        df_master = _read_file(file_paths[0])

        missing_master_keys = [k for k in key_columns if k not in df_master.columns]
        if missing_master_keys:
            raise ValueError(
                f"File Gốc '{os.path.basename(file_paths[0])}' không có cột khóa: "
                f"{', '.join(missing_master_keys)}"
            )

        _JOIN_KEY_COL = "__join_key__"
        df_master[_JOIN_KEY_COL] = _make_join_key(df_master, key_columns)

        n_total = max(1, len(file_paths) - 1)
        _report(f"✅  File Gốc: {len(df_master)} dòng. Chuẩn bị gộp {len(file_paths) - 1} file phụ...", 15)

        for i, aux_fp in enumerate(file_paths[1:], start=1):
            file_label = f"File {i + 1} ({os.path.basename(aux_fp)})"
            pct_start = 15 + int((i - 1) / n_total * 75)
            pct_end = 15 + int(i / n_total * 75)
            _report(f"🔗  Đang gộp {file_label}...", pct_start)

            try:
                df_aux = _read_file(aux_fp)
            except Exception as read_err:
                _report(f"⚠️  Bỏ qua {file_label} (lỗi đọc: {read_err})", pct_start + 2)
                continue

            missing_aux_keys = [k for k in key_columns if k not in df_aux.columns]
            if missing_aux_keys:
                _report(
                    f"⚠️  Bỏ qua {file_label} – thiếu cột khóa: {', '.join(missing_aux_keys)}",
                    pct_start + 2,
                )
                continue

            df_aux[_JOIN_KEY_COL] = _make_join_key(df_aux, key_columns)

            # --- Kiểm tra tỉ lệ khớp khóa ---
            master_keys = set(df_master[_JOIN_KEY_COL].dropna().astype(str))
            aux_keys = set(df_aux[_JOIN_KEY_COL].dropna().astype(str))
            
            import re as _re
            uuid_pattern = _re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', _re.I)
            master_valid_keys = {k for k in master_keys if not uuid_pattern.match(k)}
            aux_valid_keys = {k for k in aux_keys if not uuid_pattern.match(k)}
            
            matched = master_valid_keys.intersection(aux_valid_keys)
            if not matched and master_valid_keys and aux_valid_keys:
                raise ValueError(
                    f"File '{os.path.basename(aux_fp)}' không có dòng nào khớp khóa với File Gốc!\n\n"
                    "Nguyên nhân thường do chọn sai Cột khóa (ví dụ: cột 'Vị trí' ở File Gốc là phòng ban, "
                    "nhưng ở File Phụ lại là chức danh).\n\n"
                    "Vui lòng Quay lại và chỉ chọn 'Họ và Tên' làm khóa đối chiếu, bỏ chọn các cột không đồng nhất."
                )

            aux_data_cols = [
                c for c in df_aux.columns
                if c not in key_columns and c != _JOIN_KEY_COL
            ]

            existing_master_cols = set(df_master.columns)
            overlap_cols = [c for c in aux_data_cols if c in existing_master_cols]
            new_cols = [c for c in aux_data_cols if c not in existing_master_cols]

            df_aux_slim = df_aux[[_JOIN_KEY_COL] + aux_data_cols].copy()
            df_aux_slim = df_aux_slim.drop_duplicates(subset=[_JOIN_KEY_COL], keep="first")

            # Bước 1: Gắn cột mới (không trùng) — left join bình thường
            if new_cols:
                slim_new = df_aux_slim[[_JOIN_KEY_COL] + new_cols]
                df_master = df_master.merge(
                    slim_new, on=_JOIN_KEY_COL, how="left",
                    suffixes=("", f"_dup_{i + 1}"),
                )

            # Bước 2: Fill-priority cho cột trùng tên
            if overlap_cols:
                tmp_rename = {c: f"__aux_{c}__" for c in overlap_cols}
                slim_overlap = df_aux_slim[[_JOIN_KEY_COL] + overlap_cols].rename(columns=tmp_rename)
                df_master = df_master.merge(slim_overlap, on=_JOIN_KEY_COL, how="left")
                for c in overlap_cols:
                    tmp_col = f"__aux_{c}__"
                    if tmp_col not in df_master.columns:
                        continue
                    master_vals = df_master[c].fillna("").astype(str).str.strip()
                    aux_vals = df_master[tmp_col].fillna("").astype(str).str.strip()
                    empty_mask = master_vals.str.lower().isin({"", "nan", "none", "nat"})
                    df_master[c] = master_vals.where(~empty_mask, other=aux_vals)
                    df_master.drop(columns=[tmp_col], inplace=True)

            _report(f"✅  Đã gộp {file_label}", pct_end)

        _report("🧹  Đang dọn dẹp dữ liệu đầu ra...", 92)

        if _JOIN_KEY_COL in df_master.columns:
            df_master = df_master.drop(columns=[_JOIN_KEY_COL])

        # Xóa cột rỗng hoàn toàn TRƯỚC KHI convert NaN → None
        # (dropna không nhận Python None, chỉ nhận NaN)
        df_master = df_master.dropna(axis=1, how="all")

        df_master = df_master.where(df_master.notna(), other=None)

        for col in df_master.select_dtypes(include="object").columns:
            df_master[col] = df_master[col].apply(
                lambda v: None if str(v).strip().lower() in ("nan", "nat", "none", "") else v
            )

        # Xóa lần 2: sau cleanup string, loại cột toàn None
        empty_cols = [
            c for c in df_master.columns
            if df_master[c].apply(
                lambda v: v is None or (isinstance(v, float) and pd.isna(v))
            ).all()
        ]
        if empty_cols:
            df_master = df_master.drop(columns=empty_cols)

        df_master = df_master.reset_index(drop=True)

        _report(f"🎉  Hoàn tất! Kết quả: {len(df_master)} dòng, {len(df_master.columns)} cột.", 100)
        return df_master

    def normalize_categorical_columns(self, struct: dict, categorical_cols: List[str], mode: int, standard_columns: List[str] = None) -> "pd.DataFrame":
        """
        Dành cho Chế độ 3: Chuẩn hóa cột đánh dấu phân loại.
        mode 1: Giữ nhiều cột, chỉ sắp xếp dồn hàng.
        mode 2: Gộp thành 1 cột 'Phân loại', xóa các cột cũ.
        mode 3: Bung từ 1 cột thành nhiều cột X.
        """
        df = self.load_clean_flat_data(struct)
        
        mapping = struct.get("column_mapping", {})
        valid_cols = []
        for c in categorical_cols:
            c_norm = re.sub(r'\s+', ' ', str(c)).strip()
            mapped_name = mapping.get(c_norm) or mapping.get(c)
            if mapped_name and mapped_name != "Bỏ qua (Ignore)":
                final_name = mapped_name
            else:
                final_name = c_norm
            
            if final_name in df.columns and final_name not in valid_cols:
                valid_cols.append(final_name)
        
        if not valid_cols:
            return df
            
        if mode == 1:
            # Mode 1: Gom các dòng cùng loại lại bằng cách tạo cột rank tạm
            def get_category_rank(row):
                for i, c in enumerate(valid_cols):
                    val = str(row.get(c, "")).strip().lower()
                    if val and val not in ('nan', 'none', 'null'):
                        return i
                return 999
                
            df["_temp_rank"] = df.apply(get_category_rank, axis=1)
            # Dùng stable sort (mergesort) để giữ nguyên thứ tự phụ
            df = df.sort_values(by=["_temp_rank"], kind="mergesort").drop(columns=["_temp_rank"])
            
        elif mode == 2:
            # Mode 2: Tạo cột Phân loại và xóa cột cũ
            def get_category_name(row):
                for c in valid_cols:
                    val = str(row.get(c, "")).strip().lower()
                    if val and val not in ('nan', 'none', 'null'):
                        return c
                return "Chưa phân loại"
                
            df["Loại nhân sự"] = df.apply(get_category_name, axis=1)
            df = df.drop(columns=valid_cols)
            
            # Di chuyển cột "Loại nhân sự" lên đầu (sau cột Tên nếu có)
            cols = list(df.columns)
            cols.remove("Loại nhân sự")
            insert_idx = 0
            for i, c in enumerate(cols):
                if _norm(c) in ["ho va ten", "ho ten", "ten"]:
                    insert_idx = i + 1
                    break
            cols.insert(insert_idx, "Loại nhân sự")
            df = df[cols]
            
        elif mode == 3:
            # Mode 3: Bung từ 1 cột ra nhiều cột X (chỉ hỗ trợ khi có 1 cột được chọn)
            new_expand_cols = []
            if len(valid_cols) >= 1:
                target_col = valid_cols[0] # Lấy cột đầu tiên làm chuẩn
                # Tìm các giá trị duy nhất trong cột đó (loại bỏ rỗng)
                unique_vals = df[target_col].dropna().astype(str).str.strip().unique()
                unique_vals = [v for v in unique_vals if v and v.lower() not in ('nan', 'none', 'null')]
                
                # Tạo các cột mới dựa trên các giá trị unique và điền "x"
                for val in unique_vals:
                    df[val] = df[target_col].apply(
                        lambda x: "x" if str(x).strip() == val else ""
                    )
                
                # Xóa cột gộp ban đầu
                df = df.drop(columns=[target_col])
                new_expand_cols = unique_vals
            
        if standard_columns:
            ordered_cols = []
            # Chỉ lấy các cột CÓ TRONG df và nằm trong standard_columns
            for std_col in standard_columns:
                if std_col in df.columns:
                    ordered_cols.append(std_col)
                # Chèn Loại nhân sự ngay sau Họ và Tên nếu có
                if std_col == "Họ và Tên" and "Loại nhân sự" in df.columns:
                    ordered_cols.append("Loại nhân sự")
                # Chèn các cột mới bung ra ngay sau Họ và Tên nếu mode == 3
                if std_col == "Họ và Tên" and mode == 3 and 'new_expand_cols' in locals():
                    for exp_col in new_expand_cols:
                        if exp_col in df.columns and exp_col not in ordered_cols:
                            ordered_cols.append(exp_col)
            
            # Đảm bảo Loại nhân sự có mặt nếu Họ và Tên không có
            if "Loại nhân sự" in df.columns and "Loại nhân sự" not in ordered_cols:
                ordered_cols.append("Loại nhân sự")
                
            # Đảm bảo các cột mới bung ra có mặt nếu Họ và Tên không có
            if mode == 3 and 'new_expand_cols' in locals():
                for exp_col in new_expand_cols:
                    if exp_col in df.columns and exp_col not in ordered_cols:
                        ordered_cols.append(exp_col)
                
            # Thêm các cột còn lại (không nằm trong chuẩn)
            for col in df.columns:
                if col not in ordered_cols:
                    ordered_cols.append(col)
                    
            df = df[ordered_cols]
            
        # Thêm cột STT để đảm bảo file xuất ra luôn có STT, 
        # giúp thuật toán nhận diện dữ liệu hoạt động chính xác khi tải lại file
        if "STT" not in df.columns and "stt" not in [c.lower() for c in df.columns]:
            df.insert(0, "STT", range(1, len(df) + 1))
            
        return df